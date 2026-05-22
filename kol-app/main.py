import os
import re
import json
import math
from datetime import datetime, date
from decimal import Decimal
from typing import Optional
from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
import pymysql
from pymysql.cursors import DictCursor
import openpyxl
from io import BytesIO


def format_number(num):
    """格式化数字，如10000 -> 1万"""
    if num is None:
        return ''
    try:
        num = float(num)
        if num >= 100000000:
            return f"{num / 100000000:.1f}亿"
        elif num >= 10000:
            return f"{num / 10000:.1f}万"
        else:
            return str(int(num))
    except (ValueError, TypeError):
        return str(num)


class CustomJSONEncoder(json.JSONEncoder):
    """自定义JSON编码器，处理bytes/datetime/Decimal等特殊类型"""
    def default(self, obj):
        if isinstance(obj, bytes):
            try:
                return obj.decode('utf-8')
            except Exception:
                return str(obj)
        if isinstance(obj, datetime):
            return obj.strftime("%Y-%m-%d %H:%M:%S")
        if isinstance(obj, date):
            return obj.strftime("%Y-%m-%d")
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


app = FastAPI(title="KOL管理系统", version="2.0.0", root_path=os.environ.get("ROOT_PATH", "/kol"))

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 全局异常处理器，确保所有错误都返回JSON格式
from fastapi.exceptions import RequestValidationError
from starlette.exceptions import HTTPException as StarletteHTTPException

@app.exception_handler(StarletteHTTPException)
async def http_exception_handler(request, exc):
    return JSONResponse(
        status_code=exc.status_code,
        content={"code": exc.status_code, "message": str(exc.detail)},
    )

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    return JSONResponse(
        status_code=422,
        content={"code": 422, "message": "请求参数错误", "detail": str(exc)},
    )

@app.exception_handler(Exception)
async def general_exception_handler(request, exc):
    import traceback
    return JSONResponse(
        status_code=500,
        content={"code": 500, "message": f"服务器内部错误: {str(exc)}"},
    )

# 数据库配置
DB_CONFIG = {
    "host": os.environ.get("MYSQL_HOST", "127.0.0.1"),
    "port": int(os.environ.get("MYSQL_PORT", 3306)),
    "user": os.environ.get("MYSQL_USER", "root"),
    "password": os.environ.get("MYSQL_PASSWORD", "YOUR_DB_PASSWORD"),
    "database": os.environ.get("MYSQL_DATABASE", "YOUR_DATABASE_NAME"),
    "charset": "utf8mb4",
    "cursorclass": DictCursor,
}


def get_db():
    return pymysql.connect(**DB_CONFIG)

def _init_load_custom_fields():
    """启动时把数据库中的自定义字段加载到运行时 STANDARD_FIELDS"""
    try:
        conn = get_db()
        with conn.cursor() as cursor:
            load_custom_fields_to_standard(cursor)
        conn.close()
    except Exception as e:
        print(f"[WARN] Failed to load custom fields on startup: {e}")


def serialize_row(row):
    """将数据库行中的bytes/datetime/Decimal等类型转为可JSON序列化的格式"""
    if not row:
        return row
    for key in row:
        if isinstance(row[key], bytes):
            try:
                row[key] = row[key].decode("utf-8")
            except Exception:
                row[key] = str(row[key])
        elif isinstance(row[key], datetime):
            row[key] = row[key].strftime("%Y-%m-%d %H:%M:%S")
        elif isinstance(row[key], date):
            row[key] = row[key].strftime("%Y-%m-%d")
        elif isinstance(row[key], Decimal):
            row[key] = float(row[key])
        elif row[key] is not None and not isinstance(row[key], (str, int, float, bool, list, dict)):
            row[key] = str(row[key])
    return row


def serialize_rows(rows):
    """批量序列化数据库行"""
    return [serialize_row(row) for row in rows]


def sanitize_field_name(name):
    """将中文表头转为安全的数据库字段名，使用 field_x 格式 + 保留原名映射"""
    # 去掉首尾空格
    name = name.strip()
    if not name:
        return None
    # 转为小写英文+数字+下划线的安全名
    safe = re.sub(r'[^a-zA-Z0-9_]', '', name.replace(' ', '_').replace('-', '_'))
    if safe and not safe[0].isdigit():
        return safe.lower()
    # 中文或特殊字符，返回None，由调用方分配 field_N
    return None


def get_existing_columns(cursor):
    """获取kol_info表当前所有列名"""
    cursor.execute("SHOW COLUMNS FROM kol_info")
    return [row['Field'] for row in cursor.fetchall()]


def get_field_config(cursor):
    """获取字段配置映射: field_label -> field_key，自动去重"""
    cursor.execute("SELECT field_key, field_label, field_type FROM field_config ORDER BY sort_order, id")
    all_configs = cursor.fetchall()
    # 按 field_label 去重，保留第一条，同时统一映射为 field_name/display_name 供代码使用
    seen_display = set()
    seen_field = set()
    result = []
    for fc in all_configs:
        dn = fc['field_label']
        fn = fc['field_key']
        if dn not in seen_display and fn not in seen_field:
            seen_display.add(dn)
            seen_field.add(fn)
            result.append({
                'field_name': fn,
                'display_name': dn,
                'field_type': fc.get('field_type', 'text'),
            })
    return result


def ensure_columns(cursor, headers):
    """根据Excel表头确保数据库中有对应的列，返回 header->field_name 的映射"""
    existing_cols = get_existing_columns(cursor)
    field_configs = get_field_config(cursor)

    # 已有映射: display_name -> field_name
    display_to_field = {fc['display_name']: fc['field_name'] for fc in field_configs}
    field_to_display = {fc['field_name']: fc['display_name'] for fc in field_configs}
    existing_field_names = set(fc['field_name'] for fc in field_configs)

    # 当前最大 field_N 序号
    max_idx = 0
    for fn in existing_field_names:
        m = re.match(r'field_(\d+)', fn)
        if m:
            max_idx = max(max_idx, int(m.group(1)))

    header_mapping = {}  # excel表头 -> db字段名

    for i, h in enumerate(headers):
        h = h.strip()
        if not h:
            continue

        # 如果这个表头已经有映射了
        if h in display_to_field:
            header_mapping[h] = display_to_field[h]
            continue

        # 新表头，需要创建字段
        # 尝试用表头生成安全字段名
        safe_name = sanitize_field_name(h)
        if safe_name and safe_name not in existing_cols and safe_name not in existing_field_names and safe_name not in ('id', 'created_at', 'updated_at'):
            field_name = safe_name
        else:
            # 用 field_N 格式
            max_idx += 1
            field_name = f"field_{max_idx}"

        # 在数据库中添加列
        if field_name not in existing_cols:
            cursor.execute(f"ALTER TABLE kol_info ADD COLUMN `{field_name}` TEXT")
            existing_cols.append(field_name)

        # 记录字段配置
        cursor.execute(
            "INSERT INTO field_config (field_key, field_label, field_type, sort_order) VALUES (%s, %s, %s, %s)",
            (field_name, h, 'text', i)
        )
        display_to_field[h] = field_name
        existing_field_names.add(field_name)
        header_mapping[h] = field_name

    return header_mapping


# ============ 路由 ============

@app.get("/")
async def root():
    return RedirectResponse(url="static/index.html")


@app.get("/api/fields")
async def get_fields():
    """获取所有字段配置，前端根据此动态渲染"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            configs = get_field_config(cursor)
        return {"code": 0, "data": configs}
    finally:
        conn.close()


@app.get("/api/debug/db")
async def debug_db():
    """调试接口：查看数据库实际字段配置和数据样本"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            # 字段配置
            cursor.execute("SELECT * FROM field_config ORDER BY sort_order, id LIMIT 50")
            configs = cursor.fetchall()
            # kol_info表结构
            cursor.execute("SHOW COLUMNS FROM kol_info")
            columns = cursor.fetchall()
            # 前2条数据
            cursor.execute("SELECT * FROM kol_info LIMIT 2")
            rows = serialize_rows(cursor.fetchall())
        return {"code": 0, "data": {"field_config": configs, "columns": columns, "sample_rows": rows}}
    finally:
        conn.close()


@app.post("/api/debug/fix-fields")
async def fix_fields():
    """修复接口：清理field_config中的重复记录，保留每个display_name的最早记录"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT field_label, COUNT(*) as cnt, MIN(id) as keep_id
                FROM field_config
                GROUP BY field_label
                HAVING cnt > 1
            """)
            duplicates = cursor.fetchall()
            deleted_count = 0
            for dup in duplicates:
                cursor.execute(
                    "DELETE FROM field_config WHERE field_label = %s AND id != %s",
                    (dup['field_label'], dup['keep_id'])
                )
                deleted_count += cursor.rowcount
            conn.commit()
            cursor.execute("SELECT * FROM field_config ORDER BY sort_order, id")
            configs = cursor.fetchall()
        return {
            "code": 0,
            "message": f"修复完成，删除了 {deleted_count} 条重复记录",
            "data": {"field_config": configs, "duplicates_found": len(duplicates)}
        }
    finally:
        conn.close()


def extract_platform_name(url_or_name):
    """从URL或名称中提取平台名称"""
    if not url_or_name:
        return ''
    val = str(url_or_name).strip().lower()
    # 统一的平台名称映射（同时匹配URL和纯文本）
    platform_map = [
        (['youtube.com', 'youtu.be', 'youtube'], 'YouTube'),
        (['twitch.tv', 'twitch'], 'Twitch'),
        (['tiktok.com', 'tiktok'], 'TikTok'),
        (['instagram.com', 'instagram'], 'Instagram'),
        (['twitter.com', 'x.com', 'twitter'], 'X/Twitter'),
        (['facebook.com', 'facebook'], 'Facebook'),
        (['bilibili.com', 'bilibili'], 'Bilibili'),
        (['douyin.com', 'douyin'], '抖音'),
        (['weibo.com', 'weibo'], '微博'),
        (['xiaohongshu.com', 'xiaohongshu'], '小红书'),
    ]
    for keywords, name in platform_map:
        for kw in keywords:
            if kw in val or val == kw:
                return name
    # 不是URL，直接作为平台名（首字母大写）
    if '.' not in val and '/' not in val:
        return val.capitalize()
    return val


def normalize_genre(genre_val):
    """将genre值归类为标准名称"""
    if not genre_val:
        return '未填写'
    val = str(genre_val).strip()
    val_lower = val.lower().replace('\n', ' ').replace('\r', ' ')

    # 游戏名称归类映射（key中的关键词 -> 标准名称）
    # 注意：匹配顺序很重要，先匹配更具体的模式
    game_aliases = [
        # Call of Duty Mobile 系列
        (['codm', 'call of duty mobile', 'cod mobile'], 'Call of Duty Mobile'),
        # PUBG Mobile 系列
        (['pubg mobile', 'pubgm'], 'PUBG Mobile'),
        # Delta Force 系列
        (['delta force'], 'Delta Force'),
        # Apex 系列
        (['apex'], 'Apex'),
        # Free Fire 系列
        (['freefire', 'free fire'], 'Free Fire'),
        # Genshin 系列
        (['genshin'], 'Genshin'),
        # ARC Raiders 系列
        (['arc raiders'], 'ARC Raiders'),
        # Arena Breakout 系列
        (['arena breakout'], 'Arena Breakout'),
        # GTA 系列
        (['gta'], 'GTA'),
        # Rust 系列
        (['rust'], 'Rust'),
        # DayZ 系列
        (['dayz'], 'DayZ'),
        # Valorant 系列
        (['valorant', '瓦罗兰特'], 'Valorant'),
    ]

    # Multi/Variety 归类（包含 multi 或 variety 的）
    if 'multi' in val_lower or 'variety' in val_lower:
        return 'Multi/Variety'

    # 逐一检查游戏别名
    for keywords, standard_name in game_aliases:
        for kw in keywords:
            if kw in val_lower:
                return standard_name

    # Shooter/FPS 归类
    if val_lower in ('fps', 'shooter'):
        return 'FPS/Shooter'
    if val_lower == 'fps, survival game':
        return 'FPS/Shooter'

    # Survival 归类（纯survival或survival为主题的）
    if val_lower == 'survival' or val_lower == 'survival horror':
        return 'Survival'
    if val_lower.startswith('survival,') or val_lower.startswith('survival '):
        return 'Survival'

    # Vtuber 归类
    if 'vtuber' in val_lower:
        return 'Vtuber/Gaming'

    # 其他保持原值
    return val


@app.get("/api/dashboard")
async def get_dashboard():
    """获取仪表盘统计数据"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) as total FROM kol_info")
            total = cursor.fetchone()["total"]

            # 合作过人数：CC Pool + 往期合作记录中出现的唯一name数
            try:
                # 先从field_config找到name字段的实际列名
                configs_tmp = get_field_config(cursor)
                name_col = next((c['field_name'] for c in configs_tmp if c['display_name'] == 'name'), 'field_1')
                cursor.execute(
                    'SELECT COUNT(DISTINCT t.n) as cooperated FROM ('
                    'SELECT TRIM(`' + name_col + '`) as n FROM kol_info '
                    'WHERE is_cc = 1 AND `' + name_col + '` IS NOT NULL AND `' + name_col + "` != '' "
                    'UNION '
                    'SELECT TRIM(kol_name) as n FROM cooperation_kol_data '
                    "WHERE kol_name IS NOT NULL AND kol_name != '') t"
                )
                cooperated_count = cursor.fetchone()["cooperated"]
            except Exception:
                cooperated_count = 0

            # 获取所有字段配置
            configs = get_field_config(cursor)
            field_names = [c['field_name'] for c in configs]
            display_names = {c['field_name']: c['display_name'] for c in configs}
            existing_cols = get_existing_columns(cursor)

            # 展示3个分布图：region(国家), followers(粉丝量), genre(类型)
            # 展示3个分布图：国家分布 / 粉丝量分布 / 内容类型分布
            charts_data = []

            # 建立 display_name -> field_name 映射（如 'country' -> 'field_8'）
            display_to_col = {c['display_name']: c['field_name'] for c in configs}

            us_states_set = {
                'california', 'texas', 'florida', 'georgia', 'missouri',
                'north carolina', 'new york', 'illinois', 'pennsylvania',
                'ohio', 'michigan', 'virginia', 'washington', 'arizona',
                'massachusetts', 'tennessee', 'indiana', 'maryland',
                'wisconsin', 'colorado', 'minnesota', 'south carolina',
                'alabama', 'louisiana', 'kentucky', 'oregon', 'oklahoma',
                'connecticut', 'utah', 'iowa', 'nevada', 'arkansas',
                'mississippi', 'kansas', 'new mexico', 'nebraska',
                'idaho', 'hawaii', 'maine', 'montana', 'delaware',
                'south dakota', 'north dakota', 'alaska', 'vermont',
                'wyoming', 'west virginia', 'new hampshire', 'rhode island',
            }

            # === 1. 国家分布 ===
            try:
                col_country = display_to_col.get('country') or display_to_col.get('region') or display_to_col.get('Country')
                if col_country and col_country in existing_cols:
                    cursor.execute('SELECT `' + col_country + '` FROM kol_info WHERE `' + col_country + '` IS NOT NULL AND `' + col_country + "` != ''")
                    raw_rows = cursor.fetchall()
                    # 区域合并映射（统一大区标准）
                    na_aliases = {
                        'na', 'north america', 'united states', 'us', 'usa',
                        'canada', 'ca', 'mexico', 'north america (na)',
                    }
                    # EU：西欧
                    eu_aliases = {
                        'eu', 'europe', 'west europe', 'western europe',
                        'uk', 'united kingdom', 'united kindom', 'gb',
                        'germany', 'de', 'france', 'fr', 'spain', 'es',
                        'italy', 'it', 'netherlands', 'nl', 'sweden', 'se',
                        'norway', 'no', 'denmark', 'dk', 'finland', 'fi',
                        'switzerland', 'ch', 'austria', 'at', 'belgium', 'be',
                        'portugal', 'pt', 'ireland', 'ie',
                    }
                    # CIS：俄语区 / 东欧
                    cis_aliases = {
                        'cis', 'russia', 'russian', 'ru',
                        'east europe', 'eastern europe', 'eastern europe (ru)',
                        'ukraine', 'ua', 'belarus', 'by', 'poland', 'pl',
                        'kazakhstan', 'kz', 'turkey', 'tr', 'turkiye',
                        'romania', 'ro', 'czech', 'cz', 'hungary', 'hu',
                        'slovakia', 'sk', 'croatia', 'hr', 'serbia', 'rs',
                        'bulgaria', 'bg', 'moldova', 'md',
                    }
                    # SEA：东南亚
                    sea_aliases = {
                        'sea', 'southeast asia', 'south east asia',
                        'thailand', 'th', 'vietnam', 'vn', 'indonesia', 'id',
                        'malaysia', 'my', 'philippines', 'ph', 'singapore', 'sg',
                        'myanmar', 'mm', 'cambodia', 'kh', 'laos', 'la',
                        'brunei', 'bn', 'timor-leste',
                    }
                    # Asia：东亚 + 大洋洲 + 南亚
                    asia_aliases = {
                        'asia', 'east asia', 'aisa',  # aisa 为常见拼写错误
                        'japan', 'jp', 'korea', 'kr', 'south korea',
                        'taiwan', 'tw', 'hong kong', 'hk', 'macau', 'mo',
                        'china', 'cn', 'mongolia', 'mn',
                        'new zealand', 'nz', 'australia', 'au',
                        'india', 'in', 'south asia', 'pakistan', 'pk',
                        'bangladesh', 'bd', 'sri lanka', 'lk', 'nepal', 'np',
                    }
                    # MENA：中东北非
                    mena_aliases = {
                        'mena', 'middle east', 'arabic', 'arab',
                        'saudi arabia', 'sa', 'uae', 'ae', 'egypt', 'eg',
                        'qatar', 'qa', 'kuwait', 'kw', 'bahrain', 'bh',
                        'jordan', 'jo', 'lebanon', 'lb', 'iraq', 'iq',
                        'iran', 'ir', 'israel', 'il', 'morocco', 'ma',
                        'algeria', 'dz', 'tunisia', 'tn', 'libya', 'ly',
                    }
                    # LATAM：拉丁美洲
                    latam_aliases = {
                        'latam', 'latin america', 'south america',
                        'brazil', 'br', 'argentina', 'ar', 'colombia', 'co',
                        'columbia',  # 常见拼写错误
                        'chile', 'cl', 'peru', 'pe', 'venezuela', 've',
                        'mexico/spain',  # 特殊组合，优先归 LATAM
                    }
                    def _norm_country(v):
                        vl = v.lower().strip()
                        base = vl.split('(')[0].split('/')[0].strip()
                        # 处理 'Other*...' 等杂项
                        if base.startswith('other') or vl.startswith('other'):
                            return 'Others'
                        for aliases, label in [
                            (na_aliases, 'NA'),
                            (eu_aliases, 'EU'),
                            (cis_aliases, 'CIS'),
                            (sea_aliases, 'SEA'),
                            (asia_aliases, 'Asia'),
                            (mena_aliases, 'MENA'),
                            (latam_aliases, 'LATAM'),
                        ]:
                            if vl in aliases or base in aliases or vl in us_states_set or base in us_states_set:
                                return label
                        return v.strip()
                    country_counts = {}
                    for row in raw_rows:
                        val = str(row[col_country]).strip()
                        normalized = _norm_country(val)
                        country_counts[normalized] = country_counts.get(normalized, 0) + 1
                    stats = sorted(
                        [{'country': name, 'cnt': cnt} for name, cnt in country_counts.items()],
                        key=lambda x: x['cnt'], reverse=True
                    )[:15]
                    if stats:
                        charts_data.append({'field_name': 'country', 'display_name': '国家', 'stats': stats})
            except Exception:
                pass

            # === 2. 粉丝量分布 ===
            try:
                col_followers = display_to_col.get('followers')
                if col_followers and col_followers in existing_cols:
                    sql_f = (
                        'SELECT CASE'
                        ' WHEN CAST(`' + col_followers + '` AS UNSIGNED) >= 10000000 THEN \'1000万+\''
                        ' WHEN CAST(`' + col_followers + '` AS UNSIGNED) >= 5000000 THEN \'500万~1000万\''
                        ' WHEN CAST(`' + col_followers + '` AS UNSIGNED) >= 1000000 THEN \'100万~500万\''
                        ' WHEN CAST(`' + col_followers + '` AS UNSIGNED) >= 500000 THEN \'50万~100万\''
                        ' WHEN CAST(`' + col_followers + '` AS UNSIGNED) >= 100000 THEN \'10万~50万\''
                        ' WHEN CAST(`' + col_followers + '` AS UNSIGNED) >= 10000 THEN \'1万~10万\''
                        ' ELSE \'1万以下\' END as bucket, COUNT(*) as cnt'
                        ' FROM kol_info WHERE `' + col_followers + '` IS NOT NULL AND `' + col_followers + "` != '' AND `" + col_followers + "` REGEXP '^[0-9]+'"
                        ' GROUP BY 1 ORDER BY MIN(CAST(`' + col_followers + '` AS UNSIGNED)) DESC'
                    )
                    cursor.execute(sql_f)
                    rows = cursor.fetchall()
                    stats = [{'followers': r['bucket'], 'cnt': r['cnt']} for r in rows]
                    if stats:
                        charts_data.append({'field_name': 'followers', 'display_name': '粉丝量', 'stats': stats})
            except Exception:
                pass

            # === 3. 内容类型分布 ===
            try:
                col_genre = display_to_col.get('genre')
                if col_genre and col_genre in existing_cols:
                    cursor.execute('SELECT `' + col_genre + '` FROM kol_info WHERE `' + col_genre + '` IS NOT NULL AND `' + col_genre + "` != ''")
                    raw_rows = cursor.fetchall()
                    genre_counts = {}
                    for row in raw_rows:
                        val = str(row[col_genre]).strip()
                        normalized = normalize_genre(val)
                        genre_counts[normalized] = genre_counts.get(normalized, 0) + 1
                    stats = sorted(
                        [{'genre': name, 'cnt': cnt} for name, cnt in genre_counts.items()],
                        key=lambda x: x['cnt'], reverse=True
                    )[:10]
                    if stats:
                        charts_data.append({'field_name': 'genre', 'display_name': '内容类型', 'stats': stats})
            except Exception:
                pass

        return {
            "code": 0,
            "data": {
                "total": total,
                "cooperated_count": cooperated_count,
                "field_count": len(configs),
                "charts": charts_data,
            }
        }
    finally:
        conn.close()


@app.get("/api/kol/list")
async def list_kols(
    page: int = Query(1, ge=1),
    page_size: int = Query(12, ge=1, le=100),
    keyword: Optional[str] = "",
    sort_by: Optional[str] = "updated_at",
    sort_order: Optional[str] = "desc",
):
    """获取KOL列表（name相同的记录合并展示）
    
    合并规则：
    - price：展示所有记录的最低值~最高值区间
    - Platform：展示所有平台名（去重）
    - followers：展示所有记录的最低值~最高值区间
    - 其他字段：展示所有记录下该字段的内容（去重，最多显示前5个）
    """
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            configs = get_field_config(cursor)
            field_names = [c['field_name'] for c in configs]
            existing_cols = get_existing_columns(cursor)

            # 动态查出关键语义字段对应的实际列名（field_1~field_13）
            display_to_col = {c['display_name'].lower(): c['field_name'] for c in configs}
            col_name   = display_to_col.get('name',      'name')
            col_plat   = display_to_col.get('platform',  'platform')
            col_follow = display_to_col.get('followers', 'followers')
            col_price  = display_to_col.get('price',     'price')

            where_clauses = []
            params = []

            if keyword:
                kw_conditions = []
                kw = f"%{keyword}%"
                for fn in field_names:
                    if fn in existing_cols:
                        kw_conditions.append(f"`{fn}` LIKE %s")
                        params.append(kw)
                if kw_conditions:
                    where_clauses.append("(" + " OR ".join(kw_conditions) + ")")

            where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

            # 验证排序字段
            allowed_sorts = existing_cols
            if sort_by not in allowed_sorts:
                sort_by = "updated_at"
            order = "ASC" if sort_order == "asc" else "DESC"

            # 按实际 name 列分组合并
            merge_key = col_name
            has_name_field = merge_key in existing_cols

            if has_name_field:
                # 先查去重后的总数（按name分组）
                count_sql = f"""
                    SELECT COUNT(*) as total FROM (
                        SELECT 1 FROM kol_info WHERE {where_sql}
                        GROUP BY COALESCE(`{merge_key}`, '')
                    ) as t
                """
                cursor.execute(count_sql, params)
                total = cursor.fetchone()["total"]

                # 构建分组查询
                select_parts = []
                select_parts.append("MAX(id) as id")
                select_parts.append("COUNT(*) as merged_count")
                select_parts.append("MAX(created_at) as created_at")
                select_parts.append("MAX(updated_at) as updated_at")
                select_parts.append(f"MAX(`{col_name}`) as `name`")

                # price字段：计算MIN和MAX
                if col_price in existing_cols:
                    select_parts.append(f"MIN(CAST(`{col_price}` AS DECIMAL(20,2))) as price_min")
                    select_parts.append(f"MAX(CAST(`{col_price}` AS DECIMAL(20,2))) as price_max")
                    select_parts.append(f"MAX(`{col_price}`) as `price`")

                # followers字段：计算MIN和MAX
                if col_follow in existing_cols:
                    select_parts.append(f"MIN(CAST(`{col_follow}` AS UNSIGNED)) as followers_min")
                    select_parts.append(f"MAX(CAST(`{col_follow}` AS UNSIGNED)) as followers_max")
                    select_parts.append(f"MAX(`{col_follow}`) as `followers`")

                # platform字段：聚合所有平台
                if col_plat in existing_cols:
                    select_parts.append(f"GROUP_CONCAT(DISTINCT `{col_plat}` SEPARATOR '||') as all_platforms")
                    select_parts.append(f"MAX(`{col_plat}`) as `platform`")

                # 其他字段：取MAX（非空值优先）
                core_cols = {'id', 'created_at', 'updated_at', col_name, col_price, col_follow, col_plat}
                other_fields = [fn for fn in existing_cols if fn not in core_cols]
                for fn in other_fields:
                    select_parts.append(f"MAX(`{fn}`) as `{fn}`")

                # 排序字段映射
                order_field = f"MAX(`{sort_by}`)"

                offset = (page - 1) * page_size
                data_sql = f"""
                    SELECT {', '.join(select_parts)}
                    FROM kol_info
                    WHERE {where_sql}
                    GROUP BY COALESCE(`{merge_key}`, '')
                    ORDER BY {order_field} {order}
                    LIMIT %s OFFSET %s
                """
                cursor.execute(data_sql, params + [page_size, offset])
                rows = cursor.fetchall()
                
                # 后处理：构建区间显示
                for row in rows:
                    # 处理price区间
                    if 'price_min' in row and 'price_max' in row:
                        pmin = row.get('price_min')
                        pmax = row.get('price_max')
                        if pmin is not None and pmax is not None:
                            if float(pmin) == float(pmax):
                                row['price_display'] = str(pmin)
                            else:
                                row['price_display'] = f"{pmin}~{pmax}"
                        else:
                            row['price_display'] = row.get('price', '')
                    
                    # 处理followers区间
                    if 'followers_min' in row and 'followers_max' in row:
                        fmin = row.get('followers_min')
                        fmax = row.get('followers_max')
                        if fmin is not None and fmax is not None:
                            if int(fmin) == int(fmax):
                                row['followers_display'] = format_number(int(fmin))
                            else:
                                row['followers_display'] = f"{format_number(int(fmin))}~{format_number(int(fmax))}"
                        else:
                            row['followers_display'] = row.get('followers', '')
                    
                    # 处理platform显示
                    if 'all_platforms' in row and row['all_platforms']:
                        platforms = row['all_platforms'].split('||')
                        # 提取平台名称
                        platform_names = []
                        for p in platforms:
                            pname = extract_platform_name(p.strip())
                            if pname and pname not in platform_names:
                                platform_names.append(pname)
                        row['platform_display'] = ' / '.join(platform_names[:5])
                    else:
                        row['platform_display'] = ''
                
                rows = serialize_rows(rows)
            else:
                # 没有name字段，退回原始查询
                cursor.execute(f"SELECT COUNT(*) as total FROM kol_info WHERE {where_sql}", params)
                total = cursor.fetchone()["total"]

                offset = (page - 1) * page_size
                cursor.execute(
                    f"SELECT * FROM kol_info WHERE {where_sql} ORDER BY `{sort_by}` {order} LIMIT %s OFFSET %s",
                    params + [page_size, offset],
                )
                rows = cursor.fetchall()
                rows = serialize_rows(rows)
                for r in rows:
                    r['merged_count'] = 1
                    r['all_platforms'] = r.get('platform', '')
                    r['price_display'] = r.get('price', '')
                    r['followers_display'] = r.get('followers', '')
                    r['platform_display'] = r.get('platform', '')

        return {
            "code": 0,
            "data": {
                "list": rows,
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": math.ceil(total / page_size) if total > 0 else 1,
            }
        }
    finally:
        conn.close()


@app.get("/api/kol/{kol_id}")
async def get_kol(kol_id: int):
    """获取KOL详情，包含CC状态和往期合作效果"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            # 动态查出关键字段映射
            configs = get_field_config(cursor)
            display_to_col = {c['display_name'].lower(): c['field_name'] for c in configs}
            col_name   = display_to_col.get('name',      'name')
            col_plat   = display_to_col.get('platform',  'platform')
            col_follow = display_to_col.get('followers', 'followers')
            col_price  = display_to_col.get('price',     'price')
            col_source = display_to_col.get('source',    'source')

            cursor.execute("SELECT * FROM kol_info WHERE id=%s", (kol_id,))
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="KOL不存在")
            row = serialize_row(row)

            # 将 field_X 别名为语义字段名，方便前端统一读取
            for c in configs:
                fn = c['field_name']
                dn = c['display_name']
                if fn in row and dn not in row:
                    row[dn] = row[fn]

            # === 合并同名KOL的多条记录信息 ===
            kol_name = row.get(col_name, '') or row.get('name', '')
            if kol_name and str(kol_name).strip():
                clean_name_for_merge = str(kol_name).strip()
                existing_cols = get_existing_columns(cursor)

                # 查询所有同名记录
                cursor.execute(f"SELECT * FROM kol_info WHERE TRIM(`{col_name}`)=%s", (clean_name_for_merge,))
                all_records = serialize_rows(cursor.fetchall())
                merged_count = len(all_records)
                row['merged_count'] = merged_count

                if merged_count > 1:
                    # 合并所有平台
                    if col_plat in existing_cols:
                        all_platform_vals = []
                        for rec in all_records:
                            pval = rec.get(col_plat, '')
                            if pval and str(pval).strip():
                                all_platform_vals.append(str(pval).strip())
                        row['all_platforms'] = '||'.join(list(dict.fromkeys(all_platform_vals)))
                        platform_names = []
                        for p in all_platform_vals:
                            pname = extract_platform_name(p.strip())
                            if pname and pname not in platform_names:
                                platform_names.append(pname)
                        row['platform_display'] = ' / '.join(platform_names[:10])

                    # 合并价格范围
                    if col_price in existing_cols:
                        price_vals = []
                        for rec in all_records:
                            try:
                                pv = rec.get(col_price, '')
                                if pv and str(pv).strip():
                                    price_vals.append(float(str(pv).strip()))
                            except (ValueError, TypeError):
                                pass
                        if price_vals:
                            pmin = min(price_vals)
                            pmax = max(price_vals)
                            if pmin == pmax:
                                row['price_display'] = str(pmin)
                            else:
                                row['price_display'] = f"{pmin}~{pmax}"
                            row['price_min'] = pmin
                            row['price_max'] = pmax

                    # 合并粉丝范围
                    if col_follow in existing_cols:
                        follower_vals = []
                        for rec in all_records:
                            try:
                                fv = rec.get(col_follow, '')
                                if fv and str(fv).strip():
                                    follower_vals.append(int(float(str(fv).strip())))
                            except (ValueError, TypeError):
                                pass
                        if follower_vals:
                            fmin = min(follower_vals)
                            fmax = max(follower_vals)
                            if fmin == fmax:
                                row['followers_display'] = format_number(fmin)
                            else:
                                row['followers_display'] = f"{format_number(fmin)}~{format_number(fmax)}"
                            row['followers_min'] = fmin
                            row['followers_max'] = fmax

                    # 合并其他字段：优先取非空值
                    core_cols = {'id', 'created_at', 'updated_at', col_name, col_price, col_follow, col_plat}
                    for col in existing_cols:
                        if col not in core_cols:
                            current_val = row.get(col, '')
                            if not current_val or not str(current_val).strip():
                                for rec in all_records:
                                    alt_val = rec.get(col, '')
                                    if alt_val and str(alt_val).strip():
                                        row[col] = alt_val
                                        break

                    # 保存所有子记录的关键信息供前端展示
                    sub_records = []
                    for rec in all_records:
                        sub_records.append({
                            'id': rec.get('id'),
                            'platform': extract_platform_name(str(rec.get(col_plat, '')).strip()) if rec.get(col_plat) else '',
                            'platform_url': str(rec.get(col_plat, '')).strip() if rec.get(col_plat) else '',
                            'price': rec.get(col_price, ''),
                            'followers': rec.get(col_follow, ''),
                            'source': rec.get(col_source, ''),
                        })
                    row['sub_records'] = sub_records
                else:
                    row['merged_count'] = 1
                    row['all_platforms'] = row.get(col_plat, '')
                    row['platform_display'] = extract_platform_name(str(row.get(col_plat, '')).strip()) if row.get(col_plat) else ''
                    row['price_display'] = row.get(col_price, '')
                    row['followers_display'] = row.get(col_follow, '')
                    row['sub_records'] = [{
                        'id': row.get('id'),
                        'platform': extract_platform_name(str(row.get(col_plat, '')).strip()) if row.get(col_plat) else '',
                        'platform_url': str(row.get(col_plat, '')).strip() if row.get(col_plat) else '',
                        'price': row.get(col_price, ''),
                        'followers': row.get(col_follow, ''),
                        'source': row.get(col_source, ''),
                    }]
            else:
                row['merged_count'] = 1
                row['all_platforms'] = row.get(col_plat, '')
                row['platform_display'] = ''
                row['price_display'] = row.get('price', '')
                row['followers_display'] = row.get('followers', '')
                row['sub_records'] = [{
                    'id': row.get('id'),
                    'platform': extract_platform_name(str(row.get(col_plat, '')).strip()) if row.get(col_plat) else '',
                    'platform_url': str(row.get(col_plat, '')).strip() if row.get(col_plat) else '',
                    'price': row.get('price', ''),
                    'followers': row.get('followers', ''),
                    'source': row.get(col_source, '') if 'col_source' in dir() else '',
                }]

            # 检查该name是否为CC（直接读 kol_info.is_cc 字段）
            row['is_cc'] = bool(row.get('is_cc', 0))

            # 获取往期合作效果数据
            # 处理名称匹配：合作报告生成时可能去掉了@前缀或截取了部分名称
            cooperation_history = []
            if kol_name and str(kol_name).strip():
                clean_name = str(kol_name).strip()
                # 去掉@后的核心名称
                core_name = clean_name.lstrip('@').strip()
                
                # 构建多种可能的名称变体进行精确匹配
                name_variants = list(set(filter(None, [
                    clean_name,                           # 原始名称
                    core_name,                            # 去@版本
                    '@' + core_name if core_name else '', # 加@版本
                    core_name.split()[0] if core_name and ' ' in core_name else '',  # 第一个单词（如 "KAME Gaming" -> "KAME"）
                ])))
                
                placeholders = ','.join(['%s'] * len(name_variants))
                cursor.execute(f"""
                    SELECT ckd.*, cr.title as review_title
                    FROM cooperation_kol_data ckd
                    JOIN cooperation_reviews cr ON ckd.review_id = cr.id
                    WHERE ckd.kol_name IN ({placeholders})
                    ORDER BY ckd.created_at DESC
                """, name_variants)
                cooperation_history = serialize_rows(cursor.fetchall())
                
                # 第二层：双向LIKE模糊匹配
                if not cooperation_history and core_name:
                    cursor.execute("""
                        SELECT ckd.*, cr.title as review_title
                        FROM cooperation_kol_data ckd
                        JOIN cooperation_reviews cr ON ckd.review_id = cr.id
                        WHERE ckd.kol_name LIKE %s
                           OR %s LIKE CONCAT('%%', ckd.kol_name, '%%')
                        ORDER BY ckd.created_at DESC
                    """, (f"%{core_name}%", core_name))
                    cooperation_history = serialize_rows(cursor.fetchall())
                
                # 第三层：用名称的第一个单词做LIKE匹配（处理 "KAME Gaming" vs "KAME" 的情况）
                if not cooperation_history and core_name:
                    first_word = core_name.split()[0] if ' ' in core_name else core_name
                    if first_word and len(first_word) >= 2:
                        cursor.execute("""
                            SELECT ckd.*, cr.title as review_title
                            FROM cooperation_kol_data ckd
                            JOIN cooperation_reviews cr ON ckd.review_id = cr.id
                            WHERE ckd.kol_name LIKE %s
                               OR %s LIKE CONCAT('%%', ckd.kol_name, '%%')
                            ORDER BY ckd.created_at DESC
                        """, (f"%{first_word}%", first_word))
                        cooperation_history = serialize_rows(cursor.fetchall())
            row['cooperation_history'] = cooperation_history

        return {"code": 0, "data": row}
    finally:
        conn.close()



@app.post("/api/kol/upsert-record")
async def upsert_kol_record(data: dict):
    """新增或覆盖KOL记录：同name+source则覆盖，否则新增"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            configs = get_field_config(cursor)
            display_to_col = {c["display_name"].lower(): c["field_name"] for c in configs}
            col_name   = display_to_col.get('name',   'field_1')
            col_source = display_to_col.get('source', 'field_12')

            existing_cols = get_existing_columns(cursor)

            kol_name   = str(data.get(col_name, '') or data.get('name', '')).strip()
            kol_source = str(data.get(col_source, '') or data.get('source', '')).strip()

            # 同name+source → 覆盖
            existing_id = None
            if kol_name and kol_source:
                cursor.execute(
                    f"SELECT id FROM kol_info WHERE TRIM(`{col_name}`)=%s AND TRIM(`{col_source}`)=%s LIMIT 1",
                    (kol_name, kol_source)
                )
                row = cursor.fetchone()
                if row:
                    existing_id = row['id']

            if existing_id:
                # 覆盖已有记录
                updates, values = [], []
                for key, val in data.items():
                    if key in existing_cols and key not in ('id', 'created_at', 'updated_at'):
                        updates.append(f"`{key}`=%s")
                        values.append(str(val) if val is not None else "")
                if updates:
                    values.append(existing_id)
                    cursor.execute(f"UPDATE kol_info SET {','.join(updates)} WHERE id=%s", values)
                    conn.commit()
                return {"code": 0, "data": {"id": existing_id, "action": "updated"}, "message": "记录已更新"}
            else:
                # 新增记录
                fields, values, placeholders = [], [], []
                for key, val in data.items():
                    if key in existing_cols and key not in ('id', 'created_at', 'updated_at') and val is not None:
                        fields.append(f"`{key}`")
                        values.append(str(val) if val else "")
                        placeholders.append("%s")
                if not fields:
                    raise HTTPException(status_code=400, detail="没有有效字段")
                cursor.execute(
                    f"INSERT INTO kol_info ({','.join(fields)}) VALUES ({','.join(placeholders)})",
                    values
                )
                conn.commit()
                new_id = cursor.lastrowid
                if kol_name:
                    cursor.execute(
                        "INSERT INTO pool_registry (name, pool_type, source_table, source_id) VALUES (%s, 'kol', 'kol_info', %s)",
                        (kol_name, new_id)
                    )
                    conn.commit()
                return {"code": 0, "data": {"id": new_id, "action": "created"}, "message": "记录已新增"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.post("/api/kol/create")
async def create_kol(data: dict):
    """创建KOL（动态字段）"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            existing_cols = get_existing_columns(cursor)
            fields = []
            values = []
            placeholders = []
            for key, val in data.items():
                if key in existing_cols and key not in ('id', 'created_at', 'updated_at') and val is not None:
                    fields.append(f"`{key}`")
                    values.append(str(val) if val else "")
                    placeholders.append("%s")

            if not fields:
                raise HTTPException(status_code=400, detail="没有有效字段")

            sql = f"INSERT INTO kol_info ({','.join(fields)}) VALUES ({','.join(placeholders)})"
            cursor.execute(sql, values)
            conn.commit()
            new_id = cursor.lastrowid

            # 注册到pool_registry
            kol_name = data.get('name', '').strip()
            if kol_name:
                cursor.execute(
                    "INSERT INTO pool_registry (name, pool_type, source_table, source_id) VALUES (%s, 'kol', 'kol_info', %s)",
                    (kol_name, new_id)
                )
                conn.commit()

        return {"code": 0, "data": {"id": new_id}, "message": "创建成功"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.put("/api/kol/{kol_id}")
async def update_kol(kol_id: int, data: dict):
    """更新KOL（动态字段）"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id FROM kol_info WHERE id=%s", (kol_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="KOL不存在")

            existing_cols = get_existing_columns(cursor)
            updates = []
            values = []
            for key, val in data.items():
                if key in existing_cols and key not in ('id', 'created_at', 'updated_at'):
                    updates.append(f"`{key}`=%s")
                    values.append(str(val) if val is not None else "")

            if updates:
                values.append(kol_id)
                sql = f"UPDATE kol_info SET {','.join(updates)} WHERE id=%s"
                cursor.execute(sql, values)
                conn.commit()

        return {"code": 0, "message": "更新成功"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.delete("/api/kol/{kol_id}")
async def delete_kol(kol_id: int):
    """删除KOL"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM kol_info WHERE id=%s", (kol_id,))
            conn.commit()
        return {"code": 0, "message": "删除成功"}
    finally:
        conn.close()


# 标准字段定义及其别名（用于智能匹配）
STANDARD_FIELDS = {
    "name": {
        "display": "Name",
        "aliases": ["name", "kol name", "kol", "网红名", "名称", "kol名称", "达人名称", "influencer", "达人", "influencer id", "username", "用户名", "博主名", "creator name", "creator", "handle"],
        "description": "KOL的名称或用户名",
    },
    "platform": {
        "display": "Platform",
        "aliases": ["platform", "平台", "渠道", "社交平台", "social platform", "发布平台", "channel", "social media"],
        "description": "KOL活跃的社交媒体平台",
    },
    "followers": {
        "display": "Followers",
        "aliases": ["followers", "粉丝量级", "粉丝", "粉丝数", "fans", "subscribers", "订阅数", "粉丝量", "followers #", "粉丝数量", "subscriber count", "fan count"],
        "description": "KOL的粉丝/订阅者数量",
    },
    "views": {
        "display": "Views",
        "aliases": ["views", "视频/图文总观看量", "总观看量", "观看量", "观看", "播放量", "播放", "view count", "play count", "总观看", "总播放", "观看数", "播放数", "total views", "actual kpi", "曝光量", "impressions"],
        "description": "内容获得的观看/播放次数",
    },
    "engagement": {
        "display": "Engagement",
        "aliases": ["engagement", "总互动量", "互动量", "互动", "总互动", "interactions", "互动数", "总互动数", "点赞评论转发", "engagements"],
        "description": "点赞、评论、转发等互动总数",
    },
    "cost": {
        "display": "Cost",
        "aliases": ["cost", "price", "网红总费用", "总费用", "费用", "grand total", "kol fees", "价格", "金额", "预算", "花费", "合作费用", "KOL费用", "报价", "rate", "fee", "budget"],
        "description": "合作费用或成本金额",
    },
    "link": {
        "display": "Link",
        "aliases": ["link", "url", "发布链接及截图记录", "链接及截图", "发布链接", "截图记录", "actual releasing", "post link", "screenshot", "发布链接截图", "作品链接", "发布记录"],
        "description": "发布内容的链接或截图记录",
    },
    "country": {
        "display": "Country",
        "aliases": ["country", "region", "项目执行国家", "国家", "地区", "region/country", "所在国家", "国家/地区", "geo country", "campaign geo", "geo", "location", "market"],
        "description": "KOL所在国家或地区",
    },
    "genre": {
        "display": "Genre",
        "aliases": ["genre", "游戏", "类型", "内容类型", "game", "游戏类型", "内容领域", "category", "niche", "vertical"],
        "description": "KOL的内容类型或主要游戏",
    },
    "cpm": {
        "display": "CPM",
        "aliases": ["cpm", "千次展示成本", "CPM", "千次曝光成本", "cost per mille", "cbt1 cpm", "cbt cpm"],
        "description": "千次展示成本",
    },
    "cpa": {
        "display": "CPA",
        "aliases": ["cpa", "每次行动成本", "CPA", "转化成本", "conversion cost", "cbt1 cc", "cbt cc"],
        "description": "每次行动成本",
    },
    "source": {
        "display": "Source",
        "aliases": ["source", "来源", "数据来源", "渠道来源", "from", "channel source"],
        "description": "KOL信息的来源渠道",
    },
    "notes": {
        "display": "Notes",
        "aliases": ["notes", "备注", "说明", "comment", "备注信息", "other info", "other", "其他信息", "remark", "additional info", "others", "content", "style", "email"],
        "description": "备注或说明信息",
    },
}


def match_header_to_field(header, db_synonym_map=None):
    """智能匹配Excel表头到标准字段
    
    优先级：1. 数据库header_mapping表的同义词映射  2. STANDARD_FIELDS别名
    
    返回: (field_key, match_score) 或 (None, 0)
    """
    if not header:
        return None, 0
    
    # 清理表头（移除换行符、括号内容等）
    h_clean = header.replace('\n', ' ').replace('\r', ' ').strip()
    h_lower = h_clean.lower()
    
    # 优先查数据库的表头对照关系
    if db_synonym_map and h_lower in db_synonym_map:
        return db_synonym_map[h_lower], 100
    
    best_match = None
    best_score = 0
    
    for field_key, field_info in STANDARD_FIELDS.items():
        # 检查别名匹配
        for alias in field_info["aliases"]:
            alias_lower = alias.lower()
            # 完全匹配
            if alias_lower == h_lower:
                return field_key, 100
            # 包含匹配（表头包含别名，或别名包含表头）
            if alias_lower in h_lower:
                score = 80 + len(alias) / max(len(h_clean), 1) * 20
                if score > best_score:
                    best_match = field_key
                    best_score = score
            elif h_lower in alias_lower:
                score = 60 + len(h_clean) / max(len(alias), 1) * 20
                if score > best_score:
                    best_match = field_key
                    best_score = score
    
    return best_match, best_score


@app.post("/api/kol/preview")
async def preview_excel(file: UploadFile = File(...)):
    """预览Excel文件，自动推荐字段映射"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件(.xlsx/.xls)")

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value else ""
        headers.append(val)

    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = [str(val).strip() if val is not None else "" for val in row]
        if any(cell != "" for cell in row_data):
            all_rows.append(row_data)

    wb.close()

    # 从数据库加载表头对照关系
    db_synonym_map = {}
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            db_synonym_map = get_header_mapping_dict(cursor)
    finally:
        conn.close()

    # 生成字段映射建议
    field_mapping = []
    valid_headers = [h for h in headers if h]
    
    for idx, h in enumerate(headers):
        if not h:
            continue
        matched_field, score = match_header_to_field(h, db_synonym_map)
        field_info = STANDARD_FIELDS.get(matched_field, {}) if matched_field else {}
        field_mapping.append({
            "excel_header": h,
            "excel_index": idx,
            "suggested_field": matched_field,
            "suggested_display": field_info.get("display", ""),
            "match_score": score,
            "description": field_info.get("description", ""),
            "is_confirmed": score >= 80,  # 高分匹配自动确认
        })

    return {
        "code": 0,
        "data": {
            "filename": file.filename,
            "headers": valid_headers,
            "all_headers": headers,
            "total_rows": len(all_rows),
            "preview_rows": all_rows[:10],
            "all_rows": all_rows,
            "field_mapping": field_mapping,  # 新增：字段映射建议
            "standard_fields": STANDARD_FIELDS,  # 新增：标准字段列表供用户选择
        }
    }


def parse_number_value(val):
    """尝试解析数值，返回float或None"""
    if not val:
        return None
    val = str(val).strip()
    # 移除常见前缀后缀（$, ¥, €, 逗号等）
    val = re.sub(r'[$,，¥€£\s]', '', val)
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


@app.post("/api/kol/import")
async def import_kols(file: UploadFile = File(...), field_mapping: Optional[str] = None):
    """从Excel导入KOL数据 - 支持自定义字段映射和自动合并
    
    导入逻辑：
    - 每行数据都作为独立记录插入，保留完整历史
    - 列表查询时会按name字段合并展示
    
    Args:
        file: Excel文件
        field_mapping: JSON字符串，格式为 {"Excel表头": "标准字段名", ...}
                      例如：{"网红名": "name", "粉丝量级": "followers"}
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件(.xlsx/.xls)")

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    # 读取表头
    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value else ""
        headers.append(val)

    # 过滤空表头
    valid_headers = [h for h in headers if h]
    if not valid_headers:
        raise HTTPException(status_code=400, detail="未检测到有效的表头")

    # 解析用户确认的字段映射
    user_mapping = {}
    if field_mapping:
        try:
            user_mapping = json.loads(field_mapping)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="字段映射格式错误")

    conn = get_db()
    success_count = 0
    fail_count = 0
    errors = []
    merged_count = 0  # 合并的记录数

    try:
        with conn.cursor() as cursor:
            # 获取现有字段配置
            field_configs = get_field_config(cursor)
            display_to_field = {fc['display_name']: fc['field_name'] for fc in field_configs}
            existing_cols = get_existing_columns(cursor)
            
            # 构建最终的表头到字段名的映射
            # 规则：只映射到 STANDARD_FIELDS 的固定13个字段，匹配不上的列直接忽略，不新建字段
            header_mapping = {}

            # 从数据库加载表头同义词映射
            db_synonym_map = get_header_mapping_dict(cursor)

            for idx, h in enumerate(headers):
                h = h.strip()
                if not h:
                    continue

                # 1. 如果用户在映射界面手动指定了映射，优先使用
                if h in user_mapping:
                    target_field = user_mapping[h]
                    # 只接受 STANDARD_FIELDS 中的字段，其他忽略
                    if target_field in STANDARD_FIELDS:
                        # 找到对应的 field_key (field_1 ~ field_13)
                        std_keys = list(STANDARD_FIELDS.keys())
                        if target_field in std_keys:
                            field_idx = std_keys.index(target_field) + 1
                            header_mapping[h] = f"field_{field_idx}"
                    # 不在 STANDARD_FIELDS 中的目标字段直接忽略
                    continue

                # 2. 智能匹配到 STANDARD_FIELDS
                matched_field, score = match_header_to_field(h, db_synonym_map)
                if matched_field and score >= 60:
                    std_keys = list(STANDARD_FIELDS.keys())
                    if matched_field in std_keys:
                        field_idx = std_keys.index(matched_field) + 1
                        header_mapping[h] = f"field_{field_idx}"
                # 匹配不上或分数太低 -> 直接忽略，不新建字段

            # field_config 永远固定，不随上传变化，无需 commit 修改
            conn.commit()

            # 检查name字段是否存在
            name_field = header_mapping.get('name') or header_mapping.get('网红名') or header_mapping.get('KOL名称')
            
            # 导入数据 - 每行都插入，保留完整历史记录
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    row_data = {}
                    for col_idx, val in enumerate(row):
                        if col_idx >= len(headers):
                            break
                        h = headers[col_idx].strip()
                        if not h or h not in header_mapping:
                            continue
                        field_name = header_mapping[h]
                        cell_val = str(val).strip() if val is not None else ""
                        if cell_val and cell_val != "None":
                            # 多个原始列映射到同一标准字段时，合并写入（用 " | " 分隔）
                            # name 字段例外：只取第一个值，避免 "KOL | KOL" 重复
                            if field_name in row_data:
                                if name_field and field_name != name_field:
                                    row_data[field_name] = row_data[field_name] + " | " + cell_val
                                # name 字段已有值则跳过
                            else:
                                row_data[field_name] = cell_val

                    if not row_data:
                        continue

                    # 检查是否与已有记录完全重复（所有字段值相同）
                    is_duplicate = False
                    if name_field and name_field in row_data:
                        kol_name = row_data[name_field]
                        # 查询数据库中该name的所有记录
                        cursor.execute(f"SELECT * FROM kol_info WHERE `{name_field}` = %s", (kol_name,))
                        existing_records = cursor.fetchall()
                        
                        for existing in existing_records:
                            # 检查是否所有字段值都相同
                            all_match = True
                            for fn, new_val in row_data.items():
                                old_val = existing.get(fn)
                                if old_val:
                                    old_val = str(old_val).strip()
                                if new_val != old_val:
                                    all_match = False
                                    break
                            if all_match:
                                is_duplicate = True
                                break

                    if is_duplicate:
                        # 完全重复，跳过
                        continue

                    # 插入新记录
                    fields = [f"`{fn}`" for fn in row_data.keys()]
                    values = list(row_data.values())
                    placeholders = ["%s"] * len(fields)
                    sql = f"INSERT INTO kol_info ({','.join(fields)}) VALUES ({','.join(placeholders)})"
                    cursor.execute(sql, values)
                    new_id = cursor.lastrowid

                    # 注册到pool_registry (kol pool)
                    if name_field and name_field in row_data:
                        kol_name = row_data[name_field].strip()
                        if kol_name:
                            cursor.execute(
                                "INSERT INTO pool_registry (name, pool_type, source_table, source_id) VALUES (%s, 'kol', 'kol_info', %s)",
                                (kol_name, new_id)
                            )

                    success_count += 1

                except Exception as e:
                    fail_count += 1
                    errors.append(f"第{row_idx}行: {str(e)}")

            conn.commit()

        return {
            "code": 0,
            "data": {
                "success": success_count,
                "fail": fail_count,
                "errors": errors[:10],
                "fields_created": len(header_mapping),
                "field_mapping_used": header_mapping,
            },
            "message": f"导入完成: 成功{success_count}条, 失败{fail_count}条"
        }
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
        wb.close()


# ===== 搜索辅助：游戏别名映射 =====
GAME_ALIASES_FOR_SEARCH = [
    (['codm', 'call of duty mobile', 'cod mobile', 'cod:m'], 'Call of Duty Mobile'),
    (['pubg mobile', 'pubgm'], 'PUBG Mobile'),
    (['delta force'], 'Delta Force'),
    (['apex legends', 'apex'], 'Apex'),
    (['freefire', 'free fire', 'ff'], 'Free Fire'),
    (['genshin impact', 'genshin'], 'Genshin'),
    (['arc raiders'], 'ARC Raiders'),
    (['arena breakout'], 'Arena Breakout'),
    (['gta', 'grand theft auto'], 'GTA'),
    (['valorant', '瓦罗兰特', 'valo'], 'Valorant'),
    (['minecraft', 'mc'], 'Minecraft'),
    (['fortnite', 'fn'], 'Fortnite'),
]

def expand_keywords_with_aliases(keywords: list) -> list:
    """把关键词扩展为包含别名的关键词列表"""
    expanded = list(keywords)
    kw_lower_set = {k.lower() for k in keywords}
    for aliases, standard in GAME_ALIASES_FOR_SEARCH:
        # 检查任一别名匹配 → 把所有别名都加进搜索
        if any(a in kw_lower_set for a in aliases) or standard.lower() in kw_lower_set:
            for a in aliases:
                if a not in [e.lower() for e in expanded]:
                    expanded.append(a)
            if standard.lower() not in [e.lower() for e in expanded]:
                expanded.append(standard)
    return expanded

def parse_followers_range(keywords: list):
    """从关键词中解析 followers 范围，返回 (min, max) 或 None"""
    for kw in keywords:
        kw_lower = kw.lower().replace(' ', '')
        # 匹配: followers 200000~3000000 或 followers200000-3000000 等
        m = re.search(r'followers?[:\s]*(\d+)[~\-to]+(\d+)', kw_lower)
        if m:
            return (int(m.group(1)), int(m.group(2)))
        # 匹配单值: followers 200000
        m2 = re.search(r'followers?[:\s]*(\d+)', kw_lower)
        if m2:
            v = int(m2.group(1))
            return (v, v * 10)  # 默认范围为指定值的10倍
        # 纯数字范围: 200000~3000000
        m3 = re.search(r'^(\d{5,})[\~\-](\d{5,})$', kw_lower)
        if m3:
            return (int(m3.group(1)), int(m3.group(2)))
    return None

@app.post("/api/kol/search")
async def search_kols(req: dict):
    """智能搜索推荐KOL - 支持游戏别名模糊匹配 + followers范围匹配"""
    conn = get_db()
    try:
        keywords_str = req.get("keywords", "")
        keywords_raw = [kw.strip() for kw in keywords_str.replace("，", ",").split(",") if kw.strip()]

        # 扩展游戏别名
        keywords = expand_keywords_with_aliases(keywords_raw)

        # 解析 followers 范围
        followers_range = parse_followers_range(keywords_raw)

        with conn.cursor() as cursor:
            configs = get_field_config(cursor)
            field_names = [c['field_name'] for c in configs]
            existing_cols = get_existing_columns(cursor)
            display_map = {c['field_name']: c['display_name'] for c in configs}

            # 找出 followers 对应的字段名
            col_followers = None
            for c in configs:
                if c['display_name'].lower() == 'followers':
                    if c['field_name'] in existing_cols:
                        col_followers = c['field_name']
                    break

            # 过滤掉纯 followers 范围关键词，避免 LIKE 搜索无意义字符串
            searchable_fields = [fn for fn in field_names if fn in existing_cols]
            followers_kw_pattern = re.compile(r'followers?', re.IGNORECASE)
            text_keywords = [kw for kw in keywords if not re.search(r'^\d+$', kw) and not re.search(r'followers?[:\s]*\d', kw, re.IGNORECASE)]

            where_clauses = []
            params = []

            if text_keywords:
                kw_conditions = []
                for kw in text_keywords:
                    kw_like = f"%{kw}%"
                    field_conditions = []
                    for fn in searchable_fields:
                        field_conditions.append(f"`{fn}` LIKE %s")
                        params.append(kw_like)
                    if field_conditions:
                        kw_conditions.append("(" + " OR ".join(field_conditions) + ")")
                if kw_conditions:
                    where_clauses.append("(" + " OR ".join(kw_conditions) + ")")

            # followers 范围过滤（如果有）
            if followers_range and col_followers:
                fmin, fmax = followers_range
                where_clauses.append(
                    f"(`{col_followers}` REGEXP '^[0-9]+' AND CAST(`{col_followers}` AS UNSIGNED) BETWEEN %s AND %s)"
                )
                params.append(fmin)
                params.append(fmax)

            where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

            # 构建匹配度评分（text_keywords）
            score_parts = []
            score_params = []
            weight = 10
            for fn in searchable_fields:
                for kw in text_keywords:
                    kw_like = f"%{kw}%"
                    score_parts.append(f"(CASE WHEN `{fn}` LIKE %s THEN {max(1, weight)} ELSE 0 END)")
                    score_params.append(kw_like)
                weight = max(1, weight - 1)

            # followers 范围命中加分
            if followers_range and col_followers:
                fmin, fmax = followers_range
                score_parts.append(
                    f"(CASE WHEN `{col_followers}` REGEXP '^[0-9]+' AND CAST(`{col_followers}` AS UNSIGNED) BETWEEN %s AND %s THEN 15 ELSE 0 END)"
                )
                score_params.extend([fmin, fmax])

            score_sql = " + ".join(score_parts) if score_parts else "0"

            sql = f"""
                SELECT *, ({score_sql}) as match_score
                FROM kol_info
                WHERE {where_sql}
                ORDER BY match_score DESC
                LIMIT 20
            """

            cursor.execute(sql, score_params + params)
            rows = cursor.fetchall()
            rows = serialize_rows(rows)

            # 补充 name / followers / platform 显示
            col_name = None
            col_platform = None
            for c in configs:
                dn = c['display_name'].lower()
                if dn == 'name' and c['field_name'] in existing_cols:
                    col_name = c['field_name']
                elif dn == 'platform' and c['field_name'] in existing_cols:
                    col_platform = c['field_name']

            for row in rows:
                # 统一暴露 name / followers / platform 字段供前端直接使用
                if col_name and col_name != 'name':
                    row['name'] = row.get(col_name, '')
                if col_followers and col_followers != 'followers':
                    row['followers'] = row.get(col_followers, '')
                if col_platform and col_platform != 'platform':
                    row['platform'] = row.get(col_platform, '')

                # 生成推荐理由
                reasons = []
                for kw in keywords_raw:
                    matched = False
                    for fn in searchable_fields:
                        val = str(row.get(fn) or "")
                        if kw.lower() in val.lower():
                            dn = display_map.get(fn, fn)
                            reasons.append(f"{dn}匹配「{kw}」")
                            matched = True
                            break
                    if not matched:
                        # 检查别名匹配
                        for aliases, standard in GAME_ALIASES_FOR_SEARCH:
                            if kw.lower() in aliases or kw.lower() == standard.lower():
                                for fn in searchable_fields:
                                    val = str(row.get(fn) or "").lower()
                                    if any(a in val for a in aliases) or standard.lower() in val:
                                        dn = display_map.get(fn, fn)
                                        reasons.append(f"{dn}匹配「{kw}」")
                                        break
                                break
                # followers 范围命中
                if followers_range and col_followers:
                    fval = row.get(col_followers) or row.get('followers', '')
                    try:
                        fnum = int(str(fval).strip())
                        if followers_range[0] <= fnum <= followers_range[1]:
                            reasons.append(f"followers {fnum:,} 在范围内")
                    except Exception:
                        pass

                row["match_reasons"] = reasons if reasons else ["综合匹配"]

            # 记录搜索历史
            try:
                cursor.execute(
                    "INSERT INTO search_history (keywords, filters, result_count) VALUES (%s, %s, %s)",
                    (keywords_str, json.dumps(req, ensure_ascii=False), len(rows))
                )
                conn.commit()
            except Exception:
                pass

        return {"code": 0, "data": {"list": rows, "total": len(rows), "keywords": keywords_raw}}
    finally:
        conn.close()


@app.get("/api/options")
async def get_options():
    """获取筛选选项 - 动态从各字段中提取"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            configs = get_field_config(cursor)
            existing_cols = get_existing_columns(cursor)
            options = {}
            for fc in configs:
                fn = fc['field_name']
                dn = fc['display_name']
                if fn not in existing_cols:
                    continue
                try:
                    cursor.execute(f"SELECT DISTINCT `{fn}` as val FROM kol_info WHERE `{fn}` IS NOT NULL AND `{fn}` != '' LIMIT 50")
                    vals = []
                    for row in cursor.fetchall():
                        v = row['val']
                        if isinstance(v, bytes):
                            try:
                                v = v.decode('utf-8')
                            except Exception:
                                v = str(v)
                        if v:
                            # 拆分逗号分隔的值
                            for item in str(v).split(","):
                                item = item.strip()
                                if item and item not in vals:
                                    vals.append(item)
                    if vals:
                        options[fn] = {
                            "display_name": dn,
                            "values": sorted(vals)[:30],
                        }
                except Exception:
                    pass

        return {"code": 0, "data": options}
    finally:
        conn.close()


@app.get("/api/search/history")
async def get_search_history():
    """获取搜索历史"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM search_history ORDER BY created_at DESC LIMIT 20")
            rows = serialize_rows(cursor.fetchall())
        return {"code": 0, "data": rows}
    finally:
        conn.close()


# ============ CC (Content Creator) 创作者 API ============

def get_cc_existing_columns(cursor):
    """获取cc_info表当前所有列名"""
    cursor.execute("SHOW COLUMNS FROM cc_info")
    return [row['Field'] for row in cursor.fetchall()]


def get_cc_field_config(cursor):
    """获取CC字段配置"""
    cursor.execute("SELECT field_key, field_label, field_type FROM cc_field_config ORDER BY sort_order, id")
    all_configs = cursor.fetchall()
    seen_display = set()
    seen_field = set()
    result = []
    for fc in all_configs:
        dn = fc['field_label']
        fn = fc['field_key']
        if dn not in seen_display and fn not in seen_field:
            seen_display.add(dn)
            seen_field.add(fn)
            result.append({
                'field_name': fn,
                'display_name': dn,
                'field_type': fc.get('field_type', 'text'),
            })
    return result


@app.get("/api/cc/fields")
async def get_cc_fields():
    """获取CC字段配置"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            configs = get_cc_field_config(cursor)
        return {"code": 0, "data": configs}
    finally:
        conn.close()


@app.post("/api/cc/preview")
async def preview_cc_excel(file: UploadFile = File(...)):
    """预览CC Excel文件，复用KOL的preview逻辑"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件(.xlsx/.xls)")

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value else ""
        headers.append(val)

    all_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = [str(val).strip() if val is not None else "" for val in row]
        if any(cell != "" for cell in row_data):
            all_rows.append(row_data)

    wb.close()

    # 从数据库加载表头对照关系
    db_synonym_map = {}
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            db_synonym_map = get_header_mapping_dict(cursor)
    finally:
        conn.close()

    valid_headers = [h for h in headers if h]
    field_mapping = []
    for idx, h in enumerate(headers):
        if not h:
            continue
        matched_field, score = match_header_to_field(h, db_synonym_map)
        field_info = STANDARD_FIELDS.get(matched_field, {}) if matched_field else {}
        field_mapping.append({
            "excel_header": h,
            "excel_index": idx,
            "suggested_field": matched_field,
            "suggested_display": field_info.get("display", ""),
            "match_score": score,
            "description": field_info.get("description", ""),
            "is_confirmed": score >= 80,
        })

    return {
        "code": 0,
        "data": {
            "filename": file.filename,
            "headers": valid_headers,
            "all_headers": headers,
            "total_rows": len(all_rows),
            "preview_rows": all_rows[:10],
            "all_rows": all_rows,
            "field_mapping": field_mapping,
            "standard_fields": STANDARD_FIELDS,
        }
    }


@app.post("/api/cc/import")
async def import_ccs(file: UploadFile = File(...), field_mapping: Optional[str] = None):
    """从Excel导入CC数据 —— 写入 kol_info，source=CC，is_cc=1；同名KOL也标记 is_cc=1"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持Excel文件(.xlsx/.xls)")

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value else ""
        headers.append(val)

    valid_headers = [h for h in headers if h]
    if not valid_headers:
        raise HTTPException(status_code=400, detail="未检测到有效的表头")

    user_mapping = {}
    if field_mapping:
        try:
            user_mapping = json.loads(field_mapping)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="字段映射格式错误")

    conn = get_db()
    success_count = 0
    fail_count = 0
    errors = []
    marked_count = 0  # 已有KOL被标记为CC的数量

    try:
        with conn.cursor() as cursor:
            # 构建表头到 kol_info 字段的映射（复用 STANDARD_FIELDS）
            header_mapping = {}
            db_synonym_map = get_header_mapping_dict(cursor)

            for idx, h in enumerate(headers):
                h_stripped = h.strip()
                if not h_stripped:
                    continue
                if h_stripped in user_mapping:
                    target_field = user_mapping[h_stripped]
                    if target_field in STANDARD_FIELDS:
                        std_keys = list(STANDARD_FIELDS.keys())
                        field_idx = std_keys.index(target_field) + 1
                        header_mapping[h_stripped] = f"field_{field_idx}"
                    continue
                matched_field, score = match_header_to_field(h_stripped, db_synonym_map)
                if matched_field and score >= 60:
                    std_keys = list(STANDARD_FIELDS.keys())
                    if matched_field in std_keys:
                        field_idx = std_keys.index(matched_field) + 1
                        header_mapping[h_stripped] = f"field_{field_idx}"

            # field_config 中 source 对应的字段名
            cursor.execute("SELECT field_key FROM field_config WHERE field_label='source' LIMIT 1")
            src_row = cursor.fetchone()
            source_field = src_row['field_key'] if src_row else 'field_12'

            # field_config 中 name 对应的字段名
            cursor.execute("SELECT field_key FROM field_config WHERE field_label='name' LIMIT 1")
            name_row = cursor.fetchone()
            name_field = name_row['field_key'] if name_row else 'field_1'

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    row_data = {}
                    for col_idx, val in enumerate(row):
                        if col_idx >= len(headers):
                            break
                        h = headers[col_idx].strip()
                        if not h or h not in header_mapping:
                            continue
                        field_name = header_mapping[h]
                        cell_val = str(val).strip() if val is not None else ""
                        if cell_val and cell_val != "None":
                            if field_name in row_data:
                                # name 字段只取第一个值，不合并（避免产生 'A | A' 重复）
                                if field_name != name_field:
                                    row_data[field_name] = row_data[field_name] + " | " + cell_val
                                # 已有 name 则跳过后续同字段的列
                            else:
                                row_data[field_name] = cell_val

                    if not row_data:
                        continue

                    # 强制设置 source=CC，is_cc=1
                    row_data[source_field] = "CC"
                    cc_name = row_data.get(name_field, "").strip()

                    # 把已有同名 KOL 标记为 is_cc=1
                    if cc_name:
                        cursor.execute(
                            f"UPDATE kol_info SET is_cc=1 WHERE {name_field}=%s AND is_cc=0",
                            (cc_name,)
                        )
                        marked_count += cursor.rowcount

                    # 新增一条 CC 记录到 kol_info
                    fields = [f"`{fn}`" for fn in row_data.keys()] + ["`is_cc`"]
                    values = list(row_data.values()) + [1]
                    placeholders = ["%s"] * len(fields)
                    sql = f"INSERT INTO kol_info ({','.join(fields)}) VALUES ({','.join(placeholders)})"
                    cursor.execute(sql, values)
                    new_id = cursor.lastrowid

                    # 注册到 pool_registry（kol 类型）
                    if cc_name:
                        try:
                            cursor.execute(
                                "INSERT INTO pool_registry (name, pool_type, source_table, source_id) VALUES (%s, 'kol', 'kol_info', %s)",
                                (cc_name, new_id)
                            )
                        except Exception:
                            pass  # 重复name允许存在

                    success_count += 1
                except Exception as e:
                    fail_count += 1
                    errors.append(f"第{row_idx}行: {str(e)}")

            conn.commit()

        return {
            "code": 0,
            "data": {
                "success": success_count,
                "fail": fail_count,
                "marked_existing": marked_count,
                "errors": errors[:10],
                "field_mapping_used": header_mapping,
            },
            "message": f"导入完成: 新增{success_count}条CC记录，标记已有KOL {marked_count}条"
        }
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()
        wb.close()


@app.post("/api/cc/create")
async def create_cc(data: dict):
    """手动创建CC —— 写入 kol_info，source=CC，is_cc=1；同名KOL也标记 is_cc=1"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            # 获取 source / name 字段名
            cursor.execute("SELECT field_key FROM field_config WHERE field_label='source' LIMIT 1")
            src_row = cursor.fetchone()
            source_field = src_row['field_key'] if src_row else 'field_12'
            cursor.execute("SELECT field_key FROM field_config WHERE field_label='name' LIMIT 1")
            name_row = cursor.fetchone()
            name_field = name_row['field_key'] if name_row else 'field_1'

            kol_cols = ['field_1','field_2','field_3','field_4','field_5','field_6',
                        'field_7','field_8','field_9','field_10','field_11','field_12','field_13']
            fields = []
            values = []
            placeholders = []
            for key, val in data.items():
                if key in kol_cols and val is not None:
                    fields.append(f"`{key}`")
                    values.append(str(val) if val else "")
                    placeholders.append("%s")

            # 强制 source=CC，is_cc=1
            if f"`{source_field}`" not in fields:
                fields.append(f"`{source_field}`")
                values.append("CC")
                placeholders.append("%s")
            else:
                idx = fields.index(f"`{source_field}`")
                values[idx] = "CC"

            fields.append("`is_cc`")
            values.append(1)
            placeholders.append("%s")

            if len(fields) <= 2:
                raise HTTPException(status_code=400, detail="没有有效字段")

            sql = f"INSERT INTO kol_info ({','.join(fields)}) VALUES ({','.join(placeholders)})"
            cursor.execute(sql, values)
            new_id = cursor.lastrowid

            cc_name = data.get(name_field, '').strip() or data.get('name', '').strip()
            if cc_name:
                # 把已有同名KOL也标记为CC
                cursor.execute(
                    f"UPDATE kol_info SET is_cc=1 WHERE {name_field}=%s AND id!=%s AND is_cc=0",
                    (cc_name, new_id)
                )
                try:
                    cursor.execute(
                        "INSERT INTO pool_registry (name, pool_type, source_table, source_id) VALUES (%s, 'kol', 'kol_info', %s)",
                        (cc_name, new_id)
                    )
                except Exception:
                    pass
            conn.commit()

        return {"code": 0, "data": {"id": new_id}, "message": "创建成功"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.get("/api/pool/stats")
async def pool_stats():
    """获取三个pool的统计信息"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) as cnt FROM kol_info")
            total = cursor.fetchone()['cnt']
            cursor.execute("SELECT COUNT(*) as cnt FROM kol_info WHERE is_cc=1")
            cc_count = cursor.fetchone()['cnt']
        return {
            "code": 0,
            "data": {
                "influence_pool": total,
                "kol_pool": total,
                "cc_pool": cc_count,
            }
        }
    finally:
        conn.close()


@app.post("/api/cooperation/{review_id}/save-kol-data")
async def save_cooperation_kol_data(review_id: int, data: dict):
    """保存合作回顾中每个KOL的绩效数据到cooperation_kol_data表"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id FROM cooperation_reviews WHERE id=%s", (review_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="合作回顾不存在")

            kol_list = data.get('kol_data', [])
            # 先清除该review的旧数据
            cursor.execute("DELETE FROM cooperation_kol_data WHERE review_id=%s", (review_id,))

            inserted = 0
            for kol in kol_list:
                kol_name = kol.get('name', '').strip()
                if not kol_name:
                    continue
                cursor.execute(
                    """INSERT INTO cooperation_kol_data (review_id, kol_name, views, kol_fees, cpm, cpa)
                       VALUES (%s, %s, %s, %s, %s, %s)""",
                    (review_id, kol_name,
                     float(kol.get('views', 0) or 0),
                     float(kol.get('kol_fees', 0) or 0),
                     float(kol.get('cpm', 0) or 0),
                     float(kol.get('cpa', 0) or 0))
                )
                inserted += 1

            conn.commit()
        return {"code": 0, "message": f"已保存 {inserted} 条KOL绩效数据"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# ============ 往期合作回顾 API ============

@app.get("/api/cooperation/list")
async def list_cooperations():
    """获取所有合作回顾列表"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT cr.*, COUNT(cf.id) as file_count
                FROM cooperation_reviews cr
                LEFT JOIN cooperation_files cf ON cr.id = cf.review_id
                GROUP BY cr.id
                ORDER BY cr.updated_at DESC
            """)
            rows = serialize_rows(cursor.fetchall())
        return {"code": 0, "data": rows}
    finally:
        conn.close()


@app.post("/api/cooperation/create")
async def create_cooperation(data: dict):
    """创建合作回顾"""
    title = data.get("title", "").strip()
    description = data.get("description", "").strip()
    if not title:
        raise HTTPException(status_code=400, detail="请输入合作名称")
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT INTO cooperation_reviews (title, description) VALUES (%s, %s)",
                (title, description)
            )
            conn.commit()
            new_id = cursor.lastrowid
        return {"code": 0, "data": {"id": new_id}, "message": "创建成功"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.get("/api/cooperation/{review_id}")
async def get_cooperation(review_id: int):
    """获取合作回顾详情（含文件列表）"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM cooperation_reviews WHERE id=%s", (review_id,))
            review = cursor.fetchone()
            if not review:
                raise HTTPException(status_code=404, detail="合作回顾不存在")
            review = serialize_row(review)

            cursor.execute(
                "SELECT id, review_id, file_name, file_type, file_size, extracted_text, created_at FROM cooperation_files WHERE review_id=%s ORDER BY created_at",
                (review_id,)
            )
            files = serialize_rows(cursor.fetchall())
            review['files'] = files
        return {"code": 0, "data": review}
    finally:
        conn.close()


@app.delete("/api/cooperation/{review_id}")
async def delete_cooperation(review_id: int):
    """删除合作回顾"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM cooperation_reviews WHERE id=%s", (review_id,))
            conn.commit()
        return {"code": 0, "message": "删除成功"}
    finally:
        conn.close()


@app.post("/api/cooperation/{review_id}/upload")
async def upload_cooperation_file(review_id: int, file: UploadFile = File(...)):
    """上传合作回顾附件（PDF/Excel/图片）"""
    # 验证合作回顾存在
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id FROM cooperation_reviews WHERE id=%s", (review_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="合作回顾不存在")

        filename = file.filename or "unknown"
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        content = await file.read()
        file_size = len(content)

        # 确定文件类型
        if ext in ('pdf',):
            file_type = 'pdf'
        elif ext in ('xlsx', 'xls', 'csv'):
            file_type = 'excel'
        elif ext in ('png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'):
            file_type = 'image'
        else:
            file_type = 'other'

        # 提取文本内容
        extracted_text = ""
        try:
            if file_type == 'excel':
                extracted_text = extract_excel_text(content, ext)
            elif file_type == 'pdf':
                extracted_text = extract_pdf_text(content)
            elif file_type == 'image':
                extracted_text = f"[图片文件: {filename}]"
        except Exception as e:
            extracted_text = f"[文本提取失败: {str(e)}]"

        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT INTO cooperation_files (review_id, file_name, file_type, file_size, file_data, extracted_text) VALUES (%s, %s, %s, %s, %s, %s)",
                (review_id, filename, file_type, file_size, content, extracted_text)
            )
            conn.commit()
            file_id = cursor.lastrowid

        return {
            "code": 0,
            "data": {
                "id": file_id,
                "file_name": filename,
                "file_type": file_type,
                "file_size": file_size,
                "extracted_text": extracted_text[:500] if extracted_text else "",
            },
            "message": "上传成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.delete("/api/cooperation/file/{file_id}")
async def delete_cooperation_file(file_id: int):
    """删除合作附件"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM cooperation_files WHERE id=%s", (file_id,))
            conn.commit()
        return {"code": 0, "message": "删除成功"}
    finally:
        conn.close()


@app.get("/api/cooperation/file/{file_id}/download")
async def download_cooperation_file(file_id: int):
    """下载合作附件"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT file_name, file_type, file_size, file_data FROM cooperation_files WHERE id=%s",
                (file_id,)
            )
            row = cursor.fetchone()
        if not row or not row.get('file_data'):
            from fastapi import HTTPException
            raise HTTPException(status_code=404, detail='文件不存在')
        file_data = row['file_data']
        file_name = row['file_name']
        file_type = row['file_type'] or 'application/octet-stream'
        # MIME 映射
        mime_map = {
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xls': 'application/vnd.ms-excel',
            'pdf': 'application/pdf',
            'word': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            'pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
            'txt': 'text/plain',
        }
        content_type = mime_map.get(file_type.lower(), 'application/octet-stream')
        from urllib.parse import quote
        safe_name = quote(file_name, safe='')
        from fastapi.responses import Response
        return Response(
            content=bytes(file_data) if not isinstance(file_data, bytes) else file_data,
            media_type=content_type,
            headers={
                'Content-Disposition': f'attachment; filename*=UTF-8\'\'{safe_name}',
                'Content-Length': str(len(file_data)),
            }
        )
    finally:
        conn.close()


def extract_excel_text(content: bytes, ext: str) -> str:
    """从Excel文件中提取文本，优先读取 Influencer Marketing 表
    使用 data_only=True 读取公式缓存值（而非公式字符串）
    """
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    text_parts = []
    
    # 优先查找 Influencer Marketing 表
    target_sheet = None
    for ws in wb.worksheets:
        if 'influencer marketing' in ws.title.lower():
            target_sheet = ws
            break
    
    # 如果没找到 Influencer Marketing 表，使用第一个表
    if target_sheet is None:
        target_sheet = wb.active
    
    text_parts.append("[Sheet: " + target_sheet.title + "]")
    
    # 提取列名（第一行）- 去除单元格内换行符
    headers = []
    for cell in target_sheet[1]:
        if cell.value is not None:
            val = str(cell.value).replace('\n', ' ').replace('\r', ' ').strip()
        else:
            val = ""
        headers.append(val)
    
    # 记录列名，用特殊标记
    text_parts.append("[HEADERS] " + " || ".join(headers))
    
    # 提取数据行，过滤公式字符串（data_only未能解析时的fallback）
    for row_idx, row in enumerate(target_sheet.iter_rows(min_row=2, values_only=True)):
        if row_idx > 500:  # 限制行数
            text_parts.append("...(更多数据省略)")
            break
        cells = []
        for c in row:
            if c is None:
                cells.append("")
            else:
                s = str(c).strip()
                if s.startswith('='):
                    cells.append("")
                else:
                    cells.append(s)
        if any(cells):
            text_parts.append(" | ".join(cells))
    
    wb.close()
    return "\n".join(text_parts)


def extract_pdf_text(content: bytes) -> str:
    """从PDF中提取文本（简单方式）"""
    # 尝试用简单的文本提取
    text = ""
    try:
        # 尝试在PDF二进制中找文本流
        raw = content.decode('latin-1')
        import re as _re
        # 查找文本对象
        text_objects = _re.findall(r'\((.*?)\)', raw)
        readable = []
        for t in text_objects:
            cleaned = t.strip()
            if len(cleaned) > 2 and cleaned.isprintable():
                readable.append(cleaned)
        text = "\n".join(readable[:200])
    except Exception:
        pass
    if not text:
        text = "[PDF文件，请结合其他上传的数据文件进行分析]"
    return text


@app.post("/api/cooperation/{review_id}/generate-report")
async def generate_report(review_id: int):
    """根据上传的文件内容生成合作报告"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM cooperation_reviews WHERE id=%s", (review_id,))
            review = cursor.fetchone()
            if not review:
                raise HTTPException(status_code=404, detail="合作回顾不存在")
            review = serialize_row(review)

            cursor.execute(
                "SELECT id, file_name, file_type, file_size, extracted_text FROM cooperation_files WHERE review_id=%s ORDER BY created_at",
                (review_id,)
            )
            files = serialize_rows(cursor.fetchall())

        if not files:
            raise HTTPException(status_code=400, detail="请先上传文件再生成报告")

        # 收集所有提取的文本
        all_texts = []
        file_summary = []
        for f in files:
            file_summary.append(f"- {f['file_name']} ({f['file_type']}, {format_file_size(f['file_size'])})")
            if f.get('extracted_text'):
                all_texts.append(f"=== {f['file_name']} ===\n{f['extracted_text']}")

        combined_text = "\n\n".join(all_texts)

        # 分析数据生成报告
        report = analyze_cooperation_data(review['title'], review.get('description', ''), combined_text, file_summary, files)

        # 保存报告
        report_json = json.dumps(report, ensure_ascii=False, cls=CustomJSONEncoder)
        conn2 = get_db()
        try:
            with conn2.cursor() as cursor2:
                cursor2.execute(
                    "UPDATE cooperation_reviews SET report_content=%s, status='generated' WHERE id=%s",
                    (report_json, review_id)
                )
                # 保存每个KOL的绩效数据到cooperation_kol_data表（用于KOL详情中的"往期效果"）
                cursor2.execute("DELETE FROM cooperation_kol_data WHERE review_id=%s", (review_id,))
                for kol_detail in report.get('kol_details', []):
                    kol_name = kol_detail.get('name', '').strip()
                    if not kol_name or kol_name.startswith('_unknown_'):
                        continue
                    total_views = kol_detail.get('total_views', 0) or 0
                    total_cost = kol_detail.get('total_cost', 0) or 0
                    avg_cpm = kol_detail.get('avg_cpm', 0) or 0
                    # 计算CPA：如果有直接CPA数据使用，否则留0
                    avg_cpa = kol_detail.get('avg_cpa', 0) or 0
                    cursor2.execute(
                        """INSERT INTO cooperation_kol_data (review_id, kol_name, views, kol_fees, cpm, cpa)
                           VALUES (%s, %s, %s, %s, %s, %s)""",
                        (review_id, kol_name, total_views, total_cost, avg_cpm, avg_cpa)
                    )
                conn2.commit()
        finally:
            conn2.close()

        return {"code": 0, "data": report, "message": "报告生成成功"}
    finally:
        conn.close()


def format_file_size(size):
    """格式化文件大小"""
    if not size:
        return "0B"
    if size < 1024:
        return f"{size}B"
    elif size < 1024 * 1024:
        return f"{size/1024:.1f}KB"
    else:
        return f"{size/(1024*1024):.1f}MB"


def analyze_cooperation_data(title, description, combined_text, file_summary, files):
    """分析合作数据，生成结构化报告 - 数据总结(含粉丝量分布)、分国家汇总、KOL详情、分析文件清单
    
    有效KOL定义：同时有"发布链接及截图记录"和"费用"列内容的数据行
    """
    report = {
        "title": title,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": "",
        "file_overview": file_summary,
        "key_metrics": [],
        "followers_distribution": [],  # 粉丝量分布
        "country_summary": [],  # 分国家汇总
        "kol_details": [],  # KOL详情列表（按name合并后）
    }

    if not combined_text.strip():
        report["summary"] = f"「{title}」合作回顾报告。已上传 {len(files)} 个文件，但未能从文件中提取到有效的文本数据。建议上传Excel数据表格以获取更详细的分析。"
        return report

    # ========== 多文件段解析 ==========
    # 将combined_text按文件分段（每个上传的文件是独立的一段）
    # 文件段格式: === filename.xlsx ===\n[Sheet: ...]\n[HEADERS] ...\ndata rows...
    file_segments = []
    current_segment_lines = []
    
    for line in combined_text.split('\n'):
        if line.startswith('=== ') and line.endswith(' ==='):
            # 新文件段开始，保存上一段
            if current_segment_lines:
                file_segments.append(current_segment_lines)
            current_segment_lines = []
        else:
            current_segment_lines.append(line)
    if current_segment_lines:
        file_segments.append(current_segment_lines)
    
    # 如果没有分段标记，整体作为一段
    if not file_segments:
        file_segments = [combined_text.split('\n')]
    
    # 用于全局统计的变量
    all_headers_combined = []  # 所有文件段的表头（第一段的为主）
    all_data_rows = []  # 所有数据行（仅用于计数）
    data_rows = []  # 兼容后续代码引用
    
    # 常见列名映射（支持中英文多别名，注意Excel列名可能含换行符）
    col_mappings = {
        'link_col': ['发布链接及截图记录', '链接及截图', '发布链接', '截图记录', 'actual releasing', 'link', 'url', 'post link', 'screenshot', '发布链接截图', 'delivery link', 'deliverylink', '发布情况', '发布', 'delivered', 'delivery', 'actual video', 'vod', 'releasing link', 'actual posting date'],
        'cost_col': ['网红总费用', '总费用', '费用', 'grand total kol fees', 'grand total', 'kol fees', 'cost', 'price, usd', 'price usd', 'price (usd)', 'fee (usd)', 'price', '价格', '金额', '预算', '花费', '合作费用', 'KOL费用', 'gross cost', 'unit price'],
        'views_col': ['视频/图文总观看量', '总观看量', 'total views actual', 'video/social post total views', 'total views', 'views', '观看量', '观看', '播放量', '播放', 'view count', 'play count', '总观看', '总播放', '观看数', '播放数', 'actual kpi metrics'],
        'engagement_col': ['总互动量', 'engagement-total likes', 'engagement', '互动量', '互动', '总互动', 'interactions', '互动数', '总互动数', '点赞评论转发', 'total likes', 'total engagement'],
        'followers_col': ['粉丝量级', 'followers #', 'followers', '粉丝', '粉丝数', 'fans', 'subscribers', '订阅数', '粉丝量'],
        'cpm_col': ['cpm', '千次展示成本', 'CPM', 'cpm (usd)', 'cpm(usd)', 'cpm($)', 'cpm (rmb)', 'cost per mille', 'cost per 1000'],
        'cpa_col': ['cpa', '每次行动成本', 'CPA', '转化成本', '单次转化成本', 'cost per action', 'cost per acquisition', 'cpa (usd)', 'cpa(usd)', 'cpa($)', 'cpa (rmb)', '单次获客成本', 'cost per install', 'cpi', 'cost per kpi', 'estimated cost per kpi'],
        'name_col': ['网红名', 'influencer id', 'name', '名称', 'kol名称', '达人名称', 'kol name', 'influencer', '达人', 'KOL名称'],
        'country_col': ['项目执行国家', 'campaign geo country/district', 'country/district', 'country', '国家', '地区', 'region', 'region/country', '所在国家', '国家/地区', 'geo country', 'campaign geo'],
        'platform_col': ['platform', '平台', '渠道', 'channel', 'social platform'],
    }

    def _find_col_indices(headers):
        """根据表头查找各列的索引"""
        col_indices = {}
        for col_key, col_names in col_mappings.items():
            for idx, h in enumerate(headers):
                h_clean = h.replace('\n', ' ').replace('\r', ' ').strip()
                h_lower = h_clean.lower()
                for name in col_names:
                    if name.lower() == h_lower or name.lower() in h_lower:
                        col_indices[col_key] = idx
                        break
                if col_key in col_indices:
                    break
        # 特别处理：发布链接列（模糊匹配）
        if 'link_col' not in col_indices:
            link_col_names = ['发布链接及截图记录', '链接及截图', '发布链接', '截图记录', 'actual releasing', 'link', 'url', 'post link', 'screenshot', 'delivery link', 'deliverylink', '发布情况', '发布', 'delivered', 'delivery']
            for idx, h in enumerate(headers):
                h_clean = h.replace('\n', ' ').replace('\r', ' ').strip()
                h_lower = h_clean.lower()
                for name in link_col_names:
                    if name.lower() in h_lower:
                        col_indices['link_col'] = idx
                        break
                if 'link_col' in col_indices:
                    break
        # 特别处理：费用列（模糊匹配）
        if 'cost_col' not in col_indices:
            cost_col_names = ['网红总费用', '总费用', '费用', 'grand total', 'kol fees', 'cost', '价格', '金额', '预算', '花费']
            for idx, h in enumerate(headers):
                h_clean = h.replace('\n', ' ').replace('\r', ' ').strip()
                h_lower = h_clean.lower()
                for name in cost_col_names:
                    if name.lower() in h_lower:
                        col_indices['cost_col'] = idx
                        break
                if 'cost_col' in col_indices:
                    break
        return col_indices

    def _parse_segment(seg_lines):
        """解析单个文件段，返回 (headers, col_indices, data_rows, kol_fees_col)"""
        seg_headers = []
        seg_data_rows = []
        for line in seg_lines:
            if line.startswith('[HEADERS]'):
                headers_str = line.replace('[HEADERS]', '').strip()
                seg_headers = [h.strip() for h in headers_str.split('||')]
            elif line.startswith('[Sheet:') or not line.strip() or '...(更多数据省略)' in line:
                continue
            else:
                cells = [c.strip() for c in line.split('|')]
                if len(cells) > 1:
                    seg_data_rows.append(cells)
        # 如果没有找到HEADERS标记，尝试从第一行解析
        if not seg_headers and seg_lines:
            for line in seg_lines:
                if line.strip() and not line.startswith('[Sheet:'):
                    seg_headers = [h.strip() for h in line.split('|')]
                    break
        seg_col_indices = _find_col_indices(seg_headers) if seg_headers else {}
        # 额外查找"网红费用"列
        seg_kol_fees_col = -1
        for idx, h in enumerate(seg_headers):
            h_clean = h.replace('\n', ' ').replace('\r', ' ').strip().lower()
            if '网红费用' in h_clean or 'kol fees' in h_clean:
                seg_kol_fees_col = idx
                break
        return seg_headers, seg_col_indices, seg_data_rows, seg_kol_fees_col

    def _filter_valid_rows(seg_data_rows, seg_col_indices, seg_kol_fees_col):
        """从数据行中筛选有效KOL行（有链接或有费用）"""
        valid = []
        link_col = seg_col_indices.get('link_col', -1)
        cost_col_idx = seg_col_indices.get('cost_col', -1)
        for row in seg_data_rows:
            link_val = row[link_col] if link_col >= 0 and link_col < len(row) else ''
            cost_val = row[cost_col_idx] if cost_col_idx >= 0 and cost_col_idx < len(row) else ''
            kol_fees_val = row[seg_kol_fees_col] if seg_kol_fees_col >= 0 and seg_kol_fees_col < len(row) else ''
            link_has_content = link_val and link_val.strip() and link_val.strip() not in ['', '-', '/', 'N/A', 'n/a', '无']
            if link_val and ('http' in str(link_val).lower() or 'www.' in str(link_val).lower()):
                link_has_content = True
            cost_has_content = False
            for val in [cost_val, kol_fees_val]:
                if val and str(val).strip() and str(val).strip() not in ['', '-', '/', 'N/A', 'n/a', '无', '0', '0.00']:
                    try:
                        val_str = str(val).replace(',', '').replace('$', '').replace('¥', '').replace('€', '').strip()
                        if val_str.startswith('='):
                            continue
                        cost_num = float(val_str)
                        if cost_num > 0:
                            cost_has_content = True
                            break
                    except (ValueError, AttributeError):
                        pass
            if link_has_content or cost_has_content:
                valid.append(row)
        return valid

    # ========== 逐文件段解析，并收集有效数据 ==========
    # 每个文件段独立解析自己的表头和列索引
    parsed_segments = []  # [(headers, col_indices, valid_rows, kol_fees_col, all_rows)]
    total_valid_kols = 0
    total_lines = 0

    for seg_lines in file_segments:
        seg_headers, seg_col_indices, seg_data_rows, seg_kol_fees_col = _parse_segment(seg_lines)
        if not seg_headers and not seg_data_rows:
            continue
        # 如果该文件段有name_col，说明是KOL数据表，尝试筛选有效行
        # 如果没有link_col和cost_col，说明可能是补充数据表（如只有CPA），也需要处理
        has_kol_filter_cols = 'link_col' in seg_col_indices or 'cost_col' in seg_col_indices
        if has_kol_filter_cols:
            seg_valid_rows = _filter_valid_rows(seg_data_rows, seg_col_indices, seg_kol_fees_col)
        else:
            # 补充数据表：所有有name的行都算有效（如只含name+CPA的表）
            name_col = seg_col_indices.get('name_col', -1)
            seg_valid_rows = []
            for row in seg_data_rows:
                name_val = row[name_col] if name_col >= 0 and name_col < len(row) else ''
                if name_val and name_val.strip() and name_val.strip() not in ['', '-', 'N/A', 'n/a']:
                    seg_valid_rows.append(row)
        total_valid_kols += len(seg_valid_rows)
        total_lines += len(seg_data_rows)
        all_data_rows.extend(seg_data_rows)
        data_rows = all_data_rows  # 兼容后续代码引用
        if not all_headers_combined:
            all_headers_combined = seg_headers
        parsed_segments.append((seg_headers, seg_col_indices, seg_valid_rows, seg_kol_fees_col, seg_data_rows))
    
    # 取第一个文件段的表头用于后续兼容
    headers = all_headers_combined
    
    # ========== 核心数据提取（仅从有效KOL中提取） ==========
    views_list = []        # 观看量列表
    cost_list = []         # 成本列表
    followers_list = []    # 粉丝量列表
    engagement_list = []   # 互动量列表
    cpm_list = []          # CPM列表
    cpa_list = []          # CPA列表
    kol_names = []         # KOL名称列表
    countries = {}         # 国家分布统计

    def parse_number(val):
        """解析数字，支持 K/M/B 单位，排除N/A等无效值"""
        if not val:
            return None
        val = str(val).strip()
        # 排除无效值
        if val.upper() in ['N/A', 'NA', '-', '--', '/', '无', '无数据', '']:
            return None
        # 排除Excel公式
        if val.startswith('='):
            return None
        val = val.replace(',', '')
        # 处理 K/M/B 后缀
        suffix = ''
        if val and val[-1].upper() in 'KMB':
            suffix = val[-1].upper()
            val = val[:-1]
        try:
            num = float(val)
            if suffix == 'K':
                num *= 1000
            elif suffix == 'M':
                num *= 1000000
            elif suffix == 'B':
                num *= 1000000000
            return num if num > 0 else None
        except (ValueError, AttributeError):
            return None

    # ========== 按name列分组汇总KOL数据（遍历所有文件段） ==========
    # 使用字典按name分组，每个KOL汇总其所有记录（跨多个文件段合并）
    kol_data_by_name = {}  # {name: {views: [...], cost: [...], ...}}

    def _extract_kol_name(row, name_col_idx):
        """从行中提取KOL名称"""
        if name_col_idx >= 0 and name_col_idx < len(row):
            name_val = row[name_col_idx].strip()
            if name_val and name_val not in ['', '-', 'N/A', 'Example', 'example', '示例', 'KOL Name', 'Influencer Name']:
                if '@' in name_val:
                    cleaned = name_val.lstrip('@').strip()
                    if cleaned:
                        return cleaned
                    at_names = re.findall(r'@(\S+)', name_val)
                    if at_names:
                        return at_names[0]
                else:
                    return name_val
        return None

    # 遍历每个文件段，用该段自己的列索引提取数据
    for seg_headers, seg_col_indices, seg_valid_rows, seg_kol_fees_col, seg_all_rows in parsed_segments:
        seg_name_col = seg_col_indices.get('name_col', -1)
        seg_views_col = seg_col_indices.get('views_col', -1)
        seg_cost_col = seg_col_indices.get('cost_col', -1)
        seg_followers_col = seg_col_indices.get('followers_col', -1)
        seg_engagement_col = seg_col_indices.get('engagement_col', -1)
        seg_cpm_col = seg_col_indices.get('cpm_col', -1)
        seg_cpa_col = seg_col_indices.get('cpa_col', -1)
        seg_link_col = seg_col_indices.get('link_col', -1)
        seg_country_col = seg_col_indices.get('country_col', -1)

        # 名称别名映射：将已知的变体名统一归一化
        # 例如 spoonkid2, spoonkid3, spoonkid4 都是 spoonkid
        # blased 是 blazed 的拼写变体
        name_alias_map = {
            'spoonkid2': 'spoonkid',
            'spoonkid3': 'spoonkid',
            'spoonkid4': 'spoonkid',
            'blased': 'blazed',
            'spookid': 'spoonkid',
            'どくきの': 'dokukino',
        }

        for row in seg_valid_rows:
            kol_name = _extract_kol_name(row, seg_name_col)
            if not kol_name:
                kol_name = f"_unknown_{len(kol_data_by_name)}"

            # 先检查名称别名映射（去空格、小写后匹配）
            kol_name_check = re.sub(r'\s+', '', kol_name.lower().strip())
            if kol_name_check in name_alias_map:
                canonical = name_alias_map[kol_name_check]
                # 找到已有的canonical key或使用标准名
                found_canonical = False
                for existing_key in kol_data_by_name:
                    if re.sub(r'\s+', '', existing_key.lower().strip()) == canonical:
                        kol_name = existing_key
                        found_canonical = True
                        break
                if not found_canonical:
                    # 首次出现，使用标准名称（首字母大写）
                    kol_name = canonical.capitalize()

            # 模糊匹配：如果精确匹配不到，尝试多种归一化策略
            # 支持: 大小写无关、空格去除(BADNESS GAMER=Badnessgamer)、包含匹配
            if kol_name not in kol_data_by_name:
                kol_name_lower = kol_name.lower().strip()
                kol_name_nospace = re.sub(r'\s+', '', kol_name_lower)  # 去除所有空格
                matched_key = None
                for existing_key in kol_data_by_name:
                    existing_lower = existing_key.lower().strip()
                    existing_nospace = re.sub(r'\s+', '', existing_lower)  # 去除所有空格
                    # 完全大小写无关匹配
                    if kol_name_lower == existing_lower:
                        matched_key = existing_key
                        break
                    # 去除空格后匹配 (BADNESS GAMER = Badnessgamer, Down to Top = DowntoTop)
                    if kol_name_nospace == existing_nospace:
                        matched_key = existing_key
                        break
                    # 包含匹配：新名称是已有名称的前缀或子串，或反之
                    if kol_name_lower in existing_lower or existing_lower in kol_name_lower:
                        matched_key = existing_key
                        break
                    # 去空格后的包含匹配
                    if kol_name_nospace in existing_nospace or existing_nospace in kol_name_nospace:
                        matched_key = existing_key
                        break
                if matched_key:
                    kol_name = matched_key

            if kol_name not in kol_data_by_name:
                kol_data_by_name[kol_name] = {
                    'views': [], 'cost': [], 'followers': [], 'engagement': [],
                    'cpm': [], 'cpa': [], 'countries': set(), 'links': [],
                    'platforms': set(), 'row_count': 0
                }

            kol_entry = kol_data_by_name[kol_name]
            kol_entry['row_count'] += 1

            # 提取 Views
            if seg_views_col >= 0 and seg_views_col < len(row):
                num = parse_number(row[seg_views_col])
                if num:
                    kol_entry['views'].append(num)

            # 提取 Cost（优先使用网红费用列）
            cost_raw = None
            if seg_kol_fees_col >= 0 and seg_kol_fees_col < len(row):
                cost_raw = row[seg_kol_fees_col]
            if not cost_raw and seg_cost_col >= 0 and seg_cost_col < len(row):
                cost_raw = row[seg_cost_col]
            if cost_raw:
                cost_clean = str(cost_raw).replace('$', '').replace('¥', '').replace('€', '').replace('USD', '').replace('CNY', '').replace(',', '').strip()
                num = parse_number(cost_clean)
                if num:
                    kol_entry['cost'].append(num)

            # 提取 Followers
            if seg_followers_col >= 0 and seg_followers_col < len(row):
                num = parse_number(row[seg_followers_col])
                if num:
                    kol_entry['followers'].append(num)

            # 提取 Engagement
            if seg_engagement_col >= 0 and seg_engagement_col < len(row):
                num = parse_number(row[seg_engagement_col])
                if num:
                    kol_entry['engagement'].append(num)

            # 提取 CPM
            if seg_cpm_col >= 0 and seg_cpm_col < len(row):
                num = parse_number(row[seg_cpm_col])
                if num:
                    kol_entry['cpm'].append(num)

            # 提取 CPA
            if seg_cpa_col >= 0 and seg_cpa_col < len(row):
                num = parse_number(row[seg_cpa_col])
                if num:
                    kol_entry['cpa'].append(num)

            # 提取发布链接
            if seg_link_col >= 0 and seg_link_col < len(row):
                link_val = row[seg_link_col].strip()
                if link_val and link_val not in ['', '-', 'N/A', 'n/a', '/']:
                    kol_entry['links'].append(link_val)
                    link_lower = link_val.lower()
                    if 'youtube' in link_lower or 'youtu.be' in link_lower:
                        kol_entry['platforms'].add('YouTube')
                    elif 'twitch' in link_lower:
                        kol_entry['platforms'].add('Twitch')
                    elif 'tiktok' in link_lower:
                        kol_entry['platforms'].add('TikTok')
                    elif 'instagram' in link_lower:
                        kol_entry['platforms'].add('Instagram')
                    elif 'twitter' in link_lower or 'x.com' in link_lower:
                        kol_entry['platforms'].add('X/Twitter')
                    elif 'bilibili' in link_lower:
                        kol_entry['platforms'].add('Bilibili')

            # 提取国家（直接使用原始值，仅做基本清理）
            if seg_country_col >= 0 and seg_country_col < len(row):
                country_val = row[seg_country_col].strip()
                if country_val and country_val not in ['', '-', 'N/A', 'n/a', '/', '无', '--']:
                    # 直接使用原始值，保留原始大小写（如 NA, SEA, EA 等区域缩写）
                    kol_entry['countries'].add(country_val)
    
    # ========== 汇总各KOL的数据（按name合并后） ==========
    # 过滤掉临时标识的KOL
    valid_kol_names = [name for name in kol_data_by_name.keys() if not name.startswith('_unknown_')]
    kol_names = valid_kol_names
    
    # 计算合并后的统计数据
    for kol_name, kol_entry in kol_data_by_name.items():
        if kol_name.startswith('_unknown_'):
            continue  # 跳过无名称的临时记录
        
        # Views: 该KOL所有记录的总观看量
        if kol_entry['views']:
            views_list.append(sum(kol_entry['views']))
        
        # Cost: 该KOL所有记录的总费用
        if kol_entry['cost']:
            cost_list.append(sum(kol_entry['cost']))
        
        # Followers: 取最大值（或平均值，通常粉丝量级相对固定）
        if kol_entry['followers']:
            followers_list.append(max(kol_entry['followers']))
        
        # Engagement: 总互动量
        if kol_entry['engagement']:
            engagement_list.append(sum(kol_entry['engagement']))
        
        # CPM: 取平均值
        if kol_entry['cpm']:
            cpm_list.extend(kol_entry['cpm'])
        
        # CPA: 取平均值
        if kol_entry['cpa']:
            cpa_list.extend(kol_entry['cpa'])
        
        # 国家分布：每个KOL只计一次
        for country_code in kol_entry['countries']:
            countries[country_code] = countries.get(country_code, 0) + 1
    
    # ========== 构建KOL详情列表（按name合并后） ==========
    for kol_name in sorted(kol_names):
        kol_entry = kol_data_by_name[kol_name]
        kol_detail = {
            "name": kol_name,
            "record_count": kol_entry['row_count'],
            "total_views": sum(kol_entry['views']) if kol_entry['views'] else 0,
            "total_cost": sum(kol_entry['cost']) if kol_entry['cost'] else 0,
            "followers": max(kol_entry['followers']) if kol_entry['followers'] else 0,
            "total_engagement": sum(kol_entry['engagement']) if kol_entry['engagement'] else 0,
            "platforms": list(kol_entry['platforms']) if kol_entry['platforms'] else [],
            "countries": list(kol_entry['countries']) if kol_entry['countries'] else [],
            "links_count": len(kol_entry['links']),
            "country": ', '.join(list(kol_entry['countries'])) if kol_entry['countries'] else '',
        }
        # 计算该KOL的平均CPM
        if kol_entry['cpm']:
            kol_detail["avg_cpm"] = sum(kol_entry['cpm']) / len(kol_entry['cpm'])
        elif kol_detail['total_views'] > 0 and kol_detail['total_cost'] > 0:
            kol_detail["avg_cpm"] = (kol_detail['total_cost'] / kol_detail['total_views']) * 1000
        # 计算该KOL的平均CPA
        if kol_entry['cpa']:
            kol_detail["avg_cpa"] = sum(kol_entry['cpa']) / len(kol_entry['cpa'])
        else:
            kol_detail["avg_cpa"] = 0
        report["kol_details"].append(kol_detail)

    # ========== 构建报告数据 ==========

    # 提前计算kol_count，后续多处使用
    kol_count = len(kol_names)  # 合并后的KOL数量

    # 1. 数据总结
    summary_parts = [f"「{title}」合作回顾分析报告"]
    if description:
        summary_parts.append(f"项目描述: {description}")
    summary_parts.append(f"基于 {len(files)} 个数据文件，Influencer Marketing 表共 {total_lines} 行数据")
    # 更新有效KOL描述，反映合并情况
    if kol_count > 0 and total_valid_kols > kol_count:
        summary_parts.append(f"筛选出 {total_valid_kols} 条有效记录，按name合并为 {kol_count} 位KOL")
    else:
        summary_parts.append(f"筛选出 {total_valid_kols} 位有效KOL（有发布链接或费用记录）")
    
    # 调试信息（当没有有效数据时，输出到summary）
    if total_valid_kols == 0:
        debug_info = []
        if not headers:
            debug_info.append("未找到表头")
        else:
            debug_info.append(f"共{len(headers)}列表头")
        if data_rows:
            debug_info.append(f"共{len(data_rows)}行数据")
        summary_parts.append("数据解析详情: " + "；".join(debug_info))

    # 2. Views - 总观看量
    total_views = 0
    avg_views = 0
    if views_list:
        total_views = sum(views_list)
        avg_views = total_views / len(views_list)
        report["key_metrics"].append({
            "label": "总观看量 (Views)",
            "value": format_large_number(total_views),
            "icon": "ri-eye-line",
            "detail": f"共{len(views_list)}个数据点，平均 {format_large_number(avg_views)}"
        })
        summary_parts.append(f"总观看量 {format_large_number(total_views)}")

    # 3. Cost - 总成本
    total_cost = 0
    avg_cost = 0
    if cost_list:
        total_cost = sum(cost_list)
        avg_cost = total_cost / len(cost_list)
        report["key_metrics"].append({
            "label": "总成本 (Cost)",
            "value": f"${total_cost:,.2f}",
            "icon": "ri-money-dollar-circle-line",
            "detail": f"共{len(cost_list)}笔费用，平均 ${avg_cost:,.2f}"
        })
        summary_parts.append(f"总成本 ${total_cost:,.2f}")

    # 4. Engagement - 总互动量
    total_engagement = 0
    avg_engagement = 0
    if engagement_list:
        total_engagement = sum(engagement_list)
        avg_engagement = total_engagement / len(engagement_list)
        report["key_metrics"].append({
            "label": "总互动量 (Engagement)",
            "value": format_large_number(total_engagement),
            "icon": "ri-heart-3-line",
            "detail": f"共{len(engagement_list)}个数据点，平均 {format_large_number(avg_engagement)}"
        })
        summary_parts.append(f"总互动量 {format_large_number(total_engagement)}")

    # 5. 平均粉丝量级
    avg_followers = 0
    if followers_list:
        avg_followers = sum(followers_list) / len(followers_list)
        max_followers = max(followers_list)
        min_followers = min(followers_list)
        report["key_metrics"].append({
            "label": "平均粉丝量级",
            "value": format_large_number(avg_followers),
            "icon": "ri-user-heart-line",
            "detail": f"范围: {format_large_number(min_followers)} ~ {format_large_number(max_followers)}"
        })

    # 6. KOL人数（按name合并后的实际KOL数量）
    total_records = sum(entry['row_count'] for name, entry in kol_data_by_name.items() if not name.startswith('_unknown_')) if kol_data_by_name else total_valid_kols
    if kol_count > 0:
        merged_info = f"合并后 {kol_count} 位KOL" if total_records > kol_count else f"{kol_count} 位KOL"
        report["key_metrics"].append({
            "label": "KOL人数",
            "value": str(kol_count),
            "icon": "ri-user-star-line",
            "detail": f"{merged_info}（原{total_valid_kols}条记录按name合并）"
        })
        summary_parts.append(f"涉及 {kol_count} 位KOL（{total_valid_kols}条记录）")
    else:
        report["key_metrics"].append({
            "label": "KOL人数",
            "value": "-",
            "icon": "ri-user-star-line",
            "detail": "未找到符合条件的有效KOL数据"
        })

    # 7. 粉丝量分布
    if followers_list:
        mega = [f for f in followers_list if f >= 1000000]
        large = [f for f in followers_list if 500000 <= f < 1000000]
        medium = [f for f in followers_list if 100000 <= f < 500000]
        small = [f for f in followers_list if f < 100000]
        report["followers_distribution"] = [
            {"tier": "百万粉 (100万+)", "count": len(mega), "percent": f"{len(mega)/len(followers_list)*100:.1f}%"},
            {"tier": "50万~百万", "count": len(large), "percent": f"{len(large)/len(followers_list)*100:.1f}%"},
            {"tier": "10~50万", "count": len(medium), "percent": f"{len(medium)/len(followers_list)*100:.1f}%"},
            {"tier": "10万以下", "count": len(small), "percent": f"{len(small)/len(followers_list)*100:.1f}%"},
        ]

    # 8. 平均CPM
    avg_cpm = 0
    if cpm_list:
        avg_cpm = sum(cpm_list) / len(cpm_list)
        report["key_metrics"].append({
            "label": "平均CPM",
            "value": f"${avg_cpm:.2f}",
            "icon": "ri-line-chart-line",
            "detail": f"范围: ${min(cpm_list):.2f} ~ ${max(cpm_list):.2f}"
        })
    elif total_views > 0 and total_cost > 0:
        # 如果没有直接的CPM数据，用总成本/总观看量计算
        calculated_cpm = (total_cost / total_views) * 1000
        report["key_metrics"].append({
            "label": "平均CPM (计算值)",
            "value": f"${calculated_cpm:.2f}",
            "icon": "ri-line-chart-line",
            "detail": f"根据总成本/总观看量计算"
        })

    # 9. 平均CPA
    avg_cpa = 0
    if cpa_list:
        avg_cpa = sum(cpa_list) / len(cpa_list)
        report["key_metrics"].append({
            "label": "平均CPA",
            "value": f"${avg_cpa:.2f}",
            "icon": "ri-focus-3-line",
            "detail": f"范围: ${min(cpa_list):.2f} ~ ${max(cpa_list):.2f}"
        })
    else:
        # 尝试从kol_details中汇总CPA数据
        kol_cpa_values = [kd.get('avg_cpa', 0) for kd in report.get('kol_details', []) if kd.get('avg_cpa', 0) > 0]
        if kol_cpa_values:
            avg_cpa = sum(kol_cpa_values) / len(kol_cpa_values)
            report["key_metrics"].append({
                "label": "平均CPA",
                "value": f"${avg_cpa:.2f}",
                "icon": "ri-focus-3-line",
                "detail": f"来自 {len(kol_cpa_values)} 位KOL的CPA数据"
            })

    report["summary"] = "。".join(summary_parts) + "。"

    # ========== 分国家汇总 ==========
    # 按country字段分组，每个国家汇总与"数据总结"一致的维度
    region_data = {}  # {region: {views: [], cost: [], followers: [], engagement: [], cpm: [], cpa: [], kol_names: set()}}
    for kol_name in sorted(kol_names):
        kol_entry = kol_data_by_name[kol_name]
        # 获取该KOL的region（可能有多个国家，取第一个）
        kol_regions = list(kol_entry.get('countries', set()))
        if not kol_regions:
            kol_regions = ['未知']
        for region in kol_regions:
            if region not in region_data:
                region_data[region] = {
                    'views': [], 'cost': [], 'followers': [], 'engagement': [],
                    'cpm': [], 'cpa': [], 'kol_names': set()
                }
            rd = region_data[region]
            rd['kol_names'].add(kol_name)
            if kol_entry['views']:
                rd['views'].append(sum(kol_entry['views']))
            if kol_entry['cost']:
                rd['cost'].append(sum(kol_entry['cost']))
            if kol_entry['followers']:
                rd['followers'].append(max(kol_entry['followers']))
            if kol_entry['engagement']:
                rd['engagement'].append(sum(kol_entry['engagement']))
            if kol_entry['cpm']:
                rd['cpm'].extend(kol_entry['cpm'])
            if kol_entry['cpa']:
                rd['cpa'].extend(kol_entry['cpa'])

    for region, rd in sorted(region_data.items(), key=lambda x: len(x[1]['kol_names']), reverse=True):
        r_total_views = sum(rd['views']) if rd['views'] else 0
        r_total_cost = sum(rd['cost']) if rd['cost'] else 0
        r_total_engagement = sum(rd['engagement']) if rd['engagement'] else 0
        r_avg_followers = (sum(rd['followers']) / len(rd['followers'])) if rd['followers'] else 0
        r_avg_cpm = (sum(rd['cpm']) / len(rd['cpm'])) if rd['cpm'] else 0
        r_avg_cpa = (sum(rd['cpa']) / len(rd['cpa'])) if rd['cpa'] else 0
        # 如果没有直接CPM数据，用成本/观看量计算
        if r_avg_cpm == 0 and r_total_views > 0 and r_total_cost > 0:
            r_avg_cpm = (r_total_cost / r_total_views) * 1000
        report["country_summary"].append({
            "country": region,
            "kol_count": len(rd['kol_names']),
            "total_views": r_total_views,
            "total_cost": r_total_cost,
            "total_engagement": r_total_engagement,
            "avg_followers": r_avg_followers,
            "avg_cpm": r_avg_cpm,
            "avg_cpa": r_avg_cpa,
        })

    return report


def format_large_number(num):
    """格式化大数字，带单位"""
    if num is None or num == "":
        return "0"
    try:
        num = float(num)
    except (ValueError, TypeError):
        return str(num)
    if num <= 0:
        return "0"
    if num >= 1000000000:
        return f"{num/1000000000:.2f}B"
    elif num >= 1000000:
        return f"{num/1000000:.2f}M"
    elif num >= 10000:
        return f"{num/1000:.1f}K"
    elif num >= 1000:
        return f"{num:,.0f}"
    elif num >= 1:
        # 如果是整数或接近整数，不显示小数；否则显示两位小数
        if num == int(num):
            return f"{int(num):,}"
        else:
            return f"{num:.2f}"
    else:
        return f"{num:.2f}"


# ============ 竞品案例 API ============

@app.get("/api/competitor/list")
async def list_competitor_cases():
    """获取所有竞品案例列表"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT id, title, description, file_name, file_type, file_size, created_at, updated_at FROM competitor_cases ORDER BY created_at DESC"
            )
            rows = cursor.fetchall()
            return {"code": 0, "data": serialize_rows(rows)}
    finally:
        conn.close()


@app.post("/api/competitor/upload")
async def upload_competitor_case(file: UploadFile = File(...), title: str = "", description: str = ""):
    """上传竞品案例文档"""
    conn = get_db()
    try:
        content = await file.read()
        file_size = len(content)
        file_name = file.filename or "未命名文件"

        # 判断文件类型
        ext = file_name.rsplit('.', 1)[-1].lower() if '.' in file_name else ''
        if ext in ('pdf',):
            file_type = 'pdf'
        elif ext in ('xlsx', 'xls', 'csv'):
            file_type = 'excel'
        elif ext in ('doc', 'docx'):
            file_type = 'word'
        elif ext in ('png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'):
            file_type = 'image'
        elif ext in ('ppt', 'pptx'):
            file_type = 'ppt'
        elif ext in ('mp4', 'avi', 'mov', 'wmv'):
            file_type = 'video'
        else:
            file_type = 'other'

        # 如果没有提供标题，使用文件名
        if not title.strip():
            title = file_name.rsplit('.', 1)[0] if '.' in file_name else file_name

        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT INTO competitor_cases (title, description, file_name, file_type, file_size, file_data) VALUES (%s, %s, %s, %s, %s, %s)",
                (title.strip(), description.strip(), file_name, file_type, file_size, content)
            )
            conn.commit()
            new_id = cursor.lastrowid

        return {"code": 0, "data": {"id": new_id, "file_name": file_name, "file_type": file_type, "file_size": file_size}, "message": "上传成功"}
    finally:
        conn.close()


@app.get("/api/competitor/{case_id}")
async def get_competitor_case(case_id: int):
    """获取竞品案例详情（不含文件数据）"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT id, title, description, file_name, file_type, file_size, created_at, updated_at FROM competitor_cases WHERE id=%s",
                (case_id,)
            )
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="竞品案例不存在")
            return {"code": 0, "data": serialize_row(row)}
    finally:
        conn.close()


@app.get("/api/competitor/{case_id}/download")
async def download_competitor_file(case_id: int):
    """下载/查看竞品案例的源文件"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT file_name, file_type, file_data FROM competitor_cases WHERE id=%s",
                (case_id,)
            )
            row = cursor.fetchone()
            if not row or not row.get('file_data'):
                raise HTTPException(status_code=404, detail="文件不存在")

            file_name = row['file_name']
            file_data = row['file_data']
            ext = file_name.rsplit('.', 1)[-1].lower() if '.' in file_name else ''

            # 根据文件类型设置Content-Type
            content_type_map = {
                'pdf': 'application/pdf',
                'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'xls': 'application/vnd.ms-excel',
                'csv': 'text/csv',
                'doc': 'application/msword',
                'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'ppt': 'application/vnd.ms-powerpoint',
                'pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation',
                'png': 'image/png',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'gif': 'image/gif',
                'bmp': 'image/bmp',
                'webp': 'image/webp',
                'mp4': 'video/mp4',
            }
            content_type = content_type_map.get(ext, 'application/octet-stream')

            # 对于图片和PDF，在浏览器中直接查看；其他文件下载
            if ext in ('pdf', 'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp'):
                disposition = 'inline'
            else:
                disposition = 'attachment'

            from urllib.parse import quote
            encoded_name = quote(file_name)

            return Response(
                content=file_data,
                media_type=content_type,
                headers={
                    "Content-Disposition": f"{disposition}; filename*=UTF-8''{encoded_name}",
                    "Content-Length": str(len(file_data)),
                }
            )
    finally:
        conn.close()


@app.put("/api/competitor/{case_id}")
async def update_competitor_case(case_id: int, data: dict):
    """更新竞品案例信息（标题/描述）"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id FROM competitor_cases WHERE id=%s", (case_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="竞品案例不存在")

            updates = []
            params = []
            if 'title' in data:
                updates.append("title=%s")
                params.append(data['title'].strip())
            if 'description' in data:
                updates.append("description=%s")
                params.append(data['description'].strip())

            if updates:
                params.append(case_id)
                cursor.execute(f"UPDATE competitor_cases SET {', '.join(updates)} WHERE id=%s", params)
                conn.commit()

        return {"code": 0, "message": "更新成功"}
    finally:
        conn.close()


@app.delete("/api/competitor/{case_id}")
async def delete_competitor_case(case_id: int):
    """删除竞品案例"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM competitor_cases WHERE id=%s", (case_id,))
            conn.commit()
        return {"code": 0, "message": "删除成功"}
    finally:
        conn.close()


# ============ 表头对照关系管理 API ============

@app.get("/api/header-mapping")
async def get_header_mappings():
    """获取所有表头对照关系（DB自定义 + 内置STANDARD_FIELDS）"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM header_mapping ORDER BY sort_order, id")
            db_rows = cursor.fetchall()

        # 构建返回列表：先返回数据库中的自定义行，再补充 STANDARD_FIELDS 内置行
        db_keys = {r['system_header'] for r in db_rows}
        result = [serialize_row(r) for r in db_rows]

        # 追加内置标准字段（DB中已存在的不重复添加）
        for idx, (key, field_info) in enumerate(STANDARD_FIELDS.items()):
            if key in db_keys:
                continue  # 已在 DB 中有自定义版本，跳过
            # aliases 中去掉与 key 相同的项，作为同义词展示
            synonyms_list = [a for a in field_info["aliases"] if a.lower() != key.lower()]
            result.append({
                "id": None,           # 内置条目无 DB id
                "system_header": key,
                "display_name": field_info["display"],
                "synonyms": ", ".join(synonyms_list),
                "description": field_info.get("description", ""),
                "sort_order": idx,
                "is_builtin": True,   # 标记为内置，前端可据此区分样式
                "created_at": None,
            })

        return {"code": 0, "data": result}
    finally:
        conn.close()


@app.post("/api/header-mapping")
async def create_header_mapping(request: Request):
    """创建表头对照关系：
    - 若 system_header 是内置字段（在 STANDARD_FIELDS 中）：只补充同义词
    - 若 system_header 是新名称：新增自定义字段，同步写入 field_config / kol_info 列
    """
    data = await request.json()
    system_header = data.get('system_header', '').strip()
    display_name = data.get('display_name', '').strip()
    synonyms = data.get('synonyms', '').strip()

    if not system_header:
        raise HTTPException(status_code=400, detail="字段标识不能为空")
    if not display_name:
        display_name = system_header

    is_builtin = system_header in STANDARD_FIELDS
    is_custom_field = 0 if is_builtin else 1

    conn = get_db()
    try:
        with conn.cursor() as cursor:
            # 检查是否已存在
            cursor.execute("SELECT id FROM header_mapping WHERE system_header=%s", (system_header,))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"'{system_header}' 已存在，请直接编辑同义词")

            if is_custom_field:
                # ---- 新增自定义字段 ----
                # 1. 为 kol_info 表新增列（用 field_N 命名）
                import re as _re
                cursor.execute("SELECT field_key FROM field_config ORDER BY id DESC LIMIT 1")
                last = cursor.fetchone()
                if last and _re.match(r'field_(\d+)', last['field_key']):
                    next_idx = int(_re.match(r'field_(\d+)', last['field_key']).group(1)) + 1
                else:
                    # 统计 field_N 最大值
                    cursor.execute("SELECT field_key FROM field_config")
                    all_keys = [r['field_key'] for r in cursor.fetchall()]
                    idxs = [int(_re.match(r'field_(\d+)', k).group(1)) for k in all_keys if _re.match(r'field_(\d+)', k)]
                    next_idx = (max(idxs) + 1) if idxs else 14
                field_key = f"field_{next_idx}"

                # 检查 kol_info 是否已有该列
                cursor.execute("SHOW COLUMNS FROM kol_info LIKE %s", (field_key,))
                if not cursor.fetchone():
                    cursor.execute(f"ALTER TABLE kol_info ADD COLUMN `{field_key}` TEXT")

                # 2. 写入 field_config（让 KOL Pool 显示此列）
                cursor.execute(
                    "INSERT INTO field_config (field_key, field_label, field_type, sort_order) VALUES (%s, %s, %s, %s)",
                    (field_key, display_name, 'text', next_idx)
                )

                # 3. 动态更新运行时 STANDARD_FIELDS
                aliases_list = [system_header]
                for s in synonyms.split(','):
                    s = s.strip()
                    if s and s not in aliases_list:
                        aliases_list.append(s)
                STANDARD_FIELDS[system_header] = {
                    "display": display_name,
                    "aliases": aliases_list,
                    "description": "用户自定义字段",
                    "is_custom": True,
                    "field_key": field_key,
                }

                # 4. 写入 header_mapping
                cursor.execute(
                    "INSERT INTO header_mapping (system_header, display_name, synonyms, sort_order, is_custom_field) VALUES (%s, %s, %s, %s, %s)",
                    (system_header, display_name, synonyms, 999, 1)
                )
                conn.commit()
                return {"code": 0, "data": {"id": cursor.lastrowid, "field_key": field_key, "is_new_field": True}, "message": f"自定义字段 '{display_name}' 创建成功，已同步到 KOL Pool"}

            else:
                # ---- 仅为内置字段补充同义词 ----
                cursor.execute(
                    "INSERT INTO header_mapping (system_header, display_name, synonyms, sort_order, is_custom_field) VALUES (%s, %s, %s, %s, %s)",
                    (system_header, display_name, synonyms, 999, 0)
                )
                conn.commit()
                return {"code": 0, "data": {"id": cursor.lastrowid, "is_new_field": False}, "message": f"同义词已添加到 '{display_name}'"}
    finally:
        conn.close()


@app.put("/api/header-mapping/{mapping_id}")
async def update_header_mapping(mapping_id: int, request: Request):
    """更新表头对照关系"""
    data = await request.json()
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            updates = []
            params = []

            if 'system_header' in data:
                val = data['system_header'].strip()
                if not val:
                    raise HTTPException(status_code=400, detail="系统表头不能为空")
                # 检查唯一性
                cursor.execute("SELECT id FROM header_mapping WHERE system_header=%s AND id!=%s", (val, mapping_id))
                if cursor.fetchone():
                    raise HTTPException(status_code=400, detail=f"系统表头 '{val}' 已存在")
                updates.append("system_header=%s")
                params.append(val)

            if 'display_name' in data:
                updates.append("display_name=%s")
                params.append(data['display_name'].strip())

            if 'synonyms' in data:
                updates.append("synonyms=%s")
                params.append(data['synonyms'].strip())

            if updates:
                params.append(mapping_id)
                cursor.execute(f"UPDATE header_mapping SET {', '.join(updates)} WHERE id=%s", params)
                conn.commit()

        return {"code": 0, "message": "更新成功"}
    finally:
        conn.close()


@app.delete("/api/header-mapping/{mapping_id}")
async def delete_header_mapping(mapping_id: int):
    """删除表头对照关系"""
    conn = get_db()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM header_mapping WHERE id=%s", (mapping_id,))
            conn.commit()
        return {"code": 0, "message": "删除成功"}
    finally:
        conn.close()


def get_header_mapping_dict(cursor):
    """从数据库读取表头对照关系，返回 {同义词小写: 系统表头} 的映射字典"""
    cursor.execute("SELECT system_header, synonyms FROM header_mapping ORDER BY sort_order")
    rows = cursor.fetchall()
    mapping = {}
    for row in rows:
        sh = row['system_header']
        # 系统表头本身也加入映射
        mapping[sh.lower()] = sh
        synonyms_str = row.get('synonyms', '') or ''
        for syn in synonyms_str.split(','):
            syn = syn.strip()
            if syn:
                mapping[syn.lower()] = sh
    return mapping

def load_custom_fields_to_standard(cursor):
    """将数据库中用户自定义的字段（is_custom_field=1）动态加载到 STANDARD_FIELDS 运行时"""
    cursor.execute("SELECT system_header, display_name, synonyms FROM header_mapping WHERE is_custom_field=1")
    rows = cursor.fetchall()
    for row in rows:
        key = row['system_header']
        if key not in STANDARD_FIELDS:
            aliases = [key]
            syn_str = row.get('synonyms', '') or ''
            for s in syn_str.split(','):
                s = s.strip()
                if s and s not in aliases:
                    aliases.append(s)
            STANDARD_FIELDS[key] = {
                "display": row.get('display_name') or key,
                "aliases": aliases,
                "description": "用户自定义字段",
                "is_custom": True,
            }


# 静态文件挂载放在最后
os.makedirs("static", exist_ok=True)
os.makedirs("uploads", exist_ok=True)
app.mount("/static", StaticFiles(directory="static", html=True), name="static")
# 启动时加载自定义字段
_init_load_custom_fields()
